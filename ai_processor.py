"""
Digit 2W Insurance Commission Grid Processor — AI Edition
============================================================
Instead of hardcoded Python field-mapping logic, this version sends raw
grid rows (from the source commission Excel) to an LLM (Groq /
llama-3.1-8b-instant, free tier) in small batches. The LLM is given
the full 27 business rules as a system prompt and returns the final
structured "rule engine" output rows directly as JSON.

A post-generation Python validator then checks every returned row
against the hard constraints from the rules (Rule Code blank, Owner
Type == ALL, etc.) and reports any violations — it does NOT silently
fix them, since the point of the validator is to catch and surface
AI mistakes, not paper over them.

Environment:
    GROQ_API_KEY must be set (https://console.groq.com — free tier).

Usage:
    python ai_processor.py <input.xlsx> <output_dir> <effect_start> <effect_end> [state1,state2,...]
"""

import os
import sys
import json
import time
import re
import threading
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from groq import Groq
except ImportError:
    Groq = None

try:
    from groq import RateLimitError as _GroqRateLimitError
except ImportError:
    class _GroqRateLimitError(Exception):
        """Placeholder so `except _GroqRateLimitError` never matches when
        the groq package is missing or doesn't expose this class."""
        pass


class _RateLimiter:
    """Paces outgoing Groq calls so we stay under the org's RPM **and** TPM caps.

    The original version only paced by RPM (a fixed min-interval between
    calls). That's not the binding constraint for this pipeline: each call
    pays a large fixed cost (the ~1500+ token RULES_PROMPT system message)
    regardless of batch size, so TPM (tokens/minute) runs out long before
    RPM (requests/minute) does — that's exactly what the 429s in production
    showed (Limit 6000 TPM, hit at ~2300 tokens/call after only 2-3 calls).

    This tracks a rolling 60s window of tokens actually consumed (estimated
    pre-call, corrected post-call from the real usage Groq returns) and
    makes wait() block until BOTH the RPM interval and the TPM budget for
    the upcoming call are satisfied. A 429 can still push things out further
    via penalize().
    """

    def __init__(self, rpm, tpm):
        self.min_interval = 60.0 / max(rpm, 1)
        self.tpm = max(tpm, 1)
        self._lock = threading.Lock()
        self._next_allowed = 0.0
        self._token_events = []  # list of (timestamp, tokens)

    def _prune(self, now):
        cutoff = now - 60.0
        self._token_events = [(t, n) for t, n in self._token_events if t > cutoff]

    def wait(self, estimated_tokens=0):
        # Clamp: a single call's estimate can never be allowed to exceed the
        # TPM budget on its own — if it could, the loop below would never
        # find a state where "used + estimated_tokens <= tpm" becomes true,
        # even with a completely empty window, and would spin forever. This
        # was a real deadlock: a batch whose prompt+completion estimate came
        # out above GROQ_TPM_LIMIT would hang call_llm_batch indefinitely,
        # with no error, no retry, no log line — exactly the "stuck at 0%
        # forever" symptom seen in production. Clamping means we may still
        # get a real 429 for an oversized call, but we'll at least DO the
        # call and let the normal retry/backoff path handle it, instead of
        # hanging before ever reaching the network.
        estimated_tokens = min(estimated_tokens, int(self.tpm * 0.95))

        deadline = time.monotonic() + 90.0  # hard safety ceiling, belt-and-suspenders
        while True:
            with self._lock:
                now = time.monotonic()
                self._prune(now)
                used = sum(n for _, n in self._token_events)
                rpm_delay = self._next_allowed - now
                tpm_delay = 0.0
                if used + estimated_tokens > self.tpm:
                    # Don't just wait for the FULL window to empty — wait only
                    # until enough of the oldest reservations age out that
                    # the new call would fit. Walk events oldest-first,
                    # subtracting until there's room.
                    running = used
                    tpm_delay = 0.0
                    for t, n in self._token_events:
                        if running + estimated_tokens <= self.tpm:
                            break
                        tpm_delay = max(0.0, (t + 60.0) - now)
                        running -= n
                delay = max(rpm_delay, tpm_delay)
                if delay <= 0 or now >= deadline:
                    self._next_allowed = max(now, self._next_allowed) + self.min_interval
                    self._token_events.append((now, estimated_tokens))
                    return
            time.sleep(min(delay, 5.0) + 0.05)

    def record_actual(self, estimated_tokens, actual_tokens):
        """Correct the running token total once we know the real usage."""
        with self._lock:
            for i in range(len(self._token_events) - 1, -1, -1):
                t, n = self._token_events[i]
                if n == estimated_tokens:
                    self._token_events[i] = (t, actual_tokens)
                    return

    def penalize(self, extra_seconds):
        with self._lock:
            self._next_allowed = max(self._next_allowed, time.monotonic() + extra_seconds)


_rate_limiter = None


def _extract_retry_after(err, default=20.0):
    """Pull a Retry-After hint out of a Groq/httpx error if the response
    included one; otherwise fall back to a conservative default wait."""
    resp = getattr(err, "response", None)
    headers = getattr(resp, "headers", None) if resp is not None else None
    if headers:
        for key in ("retry-after", "Retry-After", "x-ratelimit-reset-requests"):
            val = headers.get(key)
            if val:
                try:
                    return max(float(val), 1.0)
                except (TypeError, ValueError):
                    pass
    return default

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

GROQ_MODEL = "llama-3.1-8b-instant"

# Each sheet produces a different number of OUTPUT rows per INPUT row (see
# RULES_PROMPT rule 7/8). Batch size is tuned per sheet so the JSON response
# size stays predictable: enough headroom to avoid truncation, but not so
# much that a single call burns a big chunk of the per-minute token budget.
# Tune these via env vars if you find a sweeter spot for your account's limits.
SHEET_OUTPUT_MULTIPLIER = {
    "TW 1+5": 1,
    "TW 1+1 & SATP": 2,
    "TW SAOD with Flexi Options": 4,
}
# Defaults raised from 1 row/call. At batch size 1, the fixed ~1500-token
# RULES_PROMPT system message is paid on every single call, which is what
# blew through the 6000 TPM cap after only 2-3 calls in production. Larger
# batches amortize that fixed cost over more output rows per call, cutting
# total call count (and therefore total tokens spent on system-prompt
# overhead) substantially, while staying small enough that MAX_TOKENS_CEILING
# still comfortably covers the expected output size for each sheet's
# row-multiplier.
BATCH_SIZE_BY_SHEET = {
    "TW 1+5": int(os.environ.get("BATCH_SIZE_TW1P5", "6")),
    "TW 1+1 & SATP": int(os.environ.get("BATCH_SIZE_TW1P1", "4")),
    "TW SAOD with Flexi Options": int(os.environ.get("BATCH_SIZE_SAOD", "2")),
}

DEFAULT_BATCH_SIZE = 4

# Tuned to balance two failure modes that fight each other: too LOW and
# responses get truncated mid-row (json parse errors); too HIGH and a
# single call's own estimate can exceed the account's TPM cap on its own,
# which Groq rejects outright with a 413 before pacing even matters. 580
# (up from the original 500, but down from a since-reverted 650) gives
# headroom for longer EXCLUDE/multi-fuel strings without pushing typical
# batches close to the TPM ceiling.
TOKENS_PER_OUTPUT_ROW = 580
MIN_MAX_TOKENS = 600
# 3500 (not the since-reverted 4096): the real constraint here isn't what
# the model supports (8192) but what a single call can request without
# tripping the account's TPM cap — see the tpm_safety_cap clamp in
# call_llm_batch, which is the actual fix for both truncation AND 413s.
# This ceiling is now just a secondary cap, not the primary one.
MAX_TOKENS_CEILING = int(os.environ.get("GROQ_MAX_TOKENS_CEILING", "3500"))

MAX_RETRIES = 4
RETRY_DELAY_SEC = 2

# Groq's free tier is rate-limited org-wide on BOTH requests/minute and
# tokens/minute (commonly ~30 RPM / 6000 TPM for this model on free tier,
# but check https://console.groq.com/dashboard/limits for your account's
# actual numbers — they vary by tier and change over time). TPM is the
# binding constraint here, not RPM: each call pays the full RULES_PROMPT
# system-message cost (~1500+ tokens) no matter how small the batch is.
GROQ_RPM_LIMIT = float(os.environ.get("GROQ_RPM_LIMIT", "25"))
GROQ_TPM_LIMIT = float(os.environ.get("GROQ_TPM_LIMIT", "5500"))  # a bit under the 6000 cap for headroom
_rate_limiter = _RateLimiter(GROQ_RPM_LIMIT, GROQ_TPM_LIMIT)

OUTPUT_HEADERS = [
    "Rule Code", "Rule Name *", "IC Code *", "Product Type *", "Group",
    "Rule Type *", "Cover Type", "Business Type", "Vehicle Age", "State",
    "RTO", "Vehicle Category", "Vehicle Type", "Fuel Type", "Make", "Model",
    "Owner Type", "Usage Type", "Booking Mode", "Cover Selection Type",
    "Covers", "Addon Selection Type", "Addons", "CC From", "CC To",
    "Power From", "Power To", "GVW From", "GVW To", "Carrying From",
    "Carrying To", "NCB Type", "NCB From", "NCB To", "IDV From", "IDV To",
    "OD Discount From", "OD Discount To", "Effect Start Date *",
    "Effect End Date *", "PayIn (Commision Type)", "PayIn (Reward Type)",
    "PayIn (Amount Percentage)", "PayIn (OD Amount)", "PayIn (TP Amount)",
    "POSP (Commision Type)", "POSP (Reward Type)", "POSP (Amount Percentage)",
    "POSP (OD Amount)", "POSP (TP Amount)",
]

# JSON keys the LLM must use for each output row (snake_case mirror of headers)
JSON_FIELDS = [
    "rule_code", "rule_name", "ic_code", "product_type", "group",
    "rule_type", "cover_type", "business_type", "vehicle_age", "state",
    "rto", "vehicle_category", "vehicle_type", "fuel_type", "make", "model",
    "owner_type", "usage_type", "booking_mode", "cover_selection_type",
    "covers", "addon_selection_type", "addons", "cc_from", "cc_to",
    "power_from", "power_to", "gvw_from", "gvw_to", "carrying_from",
    "carrying_to", "ncb_type", "ncb_from", "ncb_to", "idv_from", "idv_to",
    "od_discount_from", "od_discount_to", "effect_start_date",
    "effect_end_date", "payin_commission_type", "payin_reward_type",
    "payin_amount_percentage", "payin_od_amount", "payin_tp_amount",
    "posp_commission_type", "posp_reward_type", "posp_amount_percentage",
    "posp_od_amount", "posp_tp_amount",
]

# ──────────────────────────────────────────────────────────────
# The 27 business rules — embedded verbatim as the system prompt
# ──────────────────────────────────────────────────────────────

RULES_PROMPT = """You are a data transformation engine for an insurance rule-engine upload file.
You will be given raw rows from a two-wheeler (2W) insurance commission grid, plus reference
lookup data (RTO/cluster/state mappings). For EACH input row, you must produce ONE OR MORE
output rule rows (one per applicable Vehicle Type / Fuel Type / Make / CC-or-Power band
combination implied by the segment description) following these exact rules:

1. Rule code column must always be blank (null).
2. Rule name will be unique for each row you generate (use the provided naming pattern
   exactly as instructed in the batch context — never reuse a name).
3. IC code column will be the literal string "DIGIT" for all rows.
4. Product column will always be "TW" (two-wheeler only — never 4W or commercial).
5. Group column will always be blank (null).
6. Rule type column will always be "PAYINPAYOUT".
7. Cover type / Business type:
   - If the row comes from sheet "TW 1+5": Cover Type = "Comprehensive", Business Type = "new".
   - If the row comes from sheet "TW 1+1 & SATP": this sheet generates TWO output rows per
     input row — one with Cover Type = "Comprehensive", Business Type = "renew, rollover"
     (the 1+1 OD component), and one with Cover Type = "TP", Business Type = "ALL"
     (the SATP/TP component).
   - If the row comes from sheet "TW SAOD with Flexi Options": Cover Type = "SAOD",
     Business Type = "renew, rollover".
8. Vehicle age column: for TW 1+5 and TW 1+1 & SATP rows, Vehicle Age is always "ALL".
   For TW SAOD rows, Vehicle Age is the specific year (1, 2, 3, or 4) corresponding to the
   yr1/yr2/yr3/yr4 PayIn column being used — SAOD generates 4 separate rows (one per year)
   for every segment row.
9. State column: entire state name, ALL CAPS ONLY (e.g. "MAHARASHTRA", not "MH_Good" or
   "MP_Good" — those are internal cluster codes from the grid, never output them as the state).
10. RTO column: use the actual RTO codes mapped to this cluster from the "2W RTO's" sheet
    (comma-separated), or "ALL" only if the cluster has no specific RTO subset (covers the
    whole state).
11. Vehicle Type column (Bike / Scooter / ALL): derived from the segment description that is
    already present on the raw row you were given (the "Agency/PB Seg" / "Segment" field of
    THIS row) — not a separate lookup.
12. Fuel Type column: derived from that same segment field on the raw row. Petrol can NEVER be
    output alone — it must always appear combined with other fuels, e.g.
    "Petrol, LPG, Diesel, CNG" (or the cluster-specific ordering if given), unless the segment
    is explicitly Electric-only (output "Electric") or explicitly "ALL".
13. Make column: taken directly from the "Make" field already present on the raw row you were
    given (not a lookup). If the row's segment says "Others"/excludes specific brands, output
    as "EXCLUDE: BRAND1, BRAND2, ..." using the brand names given on that row.
14. Model column: "ALL" for almost every row. Only for TW 1+1 & SATP and SAOD sheets are
    specific models sometimes excluded or included — only apply an exception if the raw row
    itself states one, otherwise default to "ALL".
15. Owner Type column: always "ALL".
16. Usage Type column: always blank (null).
17. Booking Mode column: always "any".
18. Cover Selection Type column: always "na".
19. Covers column: always blank (null).
20. Addon Selection Type column: always "na".
21. Addons column: always blank (null).
22. CC From / CC To columns: derived from the segment field already present on the raw row
    (e.g. "MC <=155" -> CC To=155; "3-7 KW" is a POWER band, not CC — see rule 23), if the
    segment implies a CC range. Otherwise blank (null).
23. Power From / Power To columns: derived from the segment field already present on the raw
    row, if the segment implies an EV power band (e.g. "3-7 KW" -> Power From=3.00,
    Power To=7.00). Otherwise blank (null).
24. Any column not explicitly covered by a rule that is empty in the source/given output stays
    blank (null). NCB Type is always "na".
25. PayIn (Commission Type) is always "net" UNLESS the source cell says "MISP", in which case
    Commission Type = "od" and Amount Percentage = "22.5".
26. PayIn (Reward Type) is always "percentage".
27. PayIn (Amount Percentage) is taken from the raw row's commission field for the relevant
    output row:
    - TW 1+5: use the row's "max_cd2" field.
    - TW 1+1 & SATP: use "max_cd2_1plus1_od" for the Comprehensive/renew-rollover (OD) output
      row, and "max_cd2_satp_tp" for the TP output row.
    - TW SAOD with Flexi Options: use "max_cd2_year1"/"max_cd2_year2"/"max_cd2_year3"/
      "max_cd2_year4" for the matching Vehicle Age (1/2/3/4) output row.
    If that field's value is the string "MISP", the amount is 22.5 (and Commission Type = "od"
    per rule 25). If the value is the string "D", the amount is 0. Otherwise use the numeric
    value given directly as a percentage (e.g. 0.35 in the source means the percentage value is
    35, since these source columns are stored as decimals — multiply by 100, e.g. 0.225 -> 22.5).

Additional fixed output rules (not in the numbered list but always true):
- PayIn (OD Amount) and PayIn (TP Amount) are always "0".
- POSP (Commission Type) is always "net", POSP (Reward Type) is always "percentage".
- POSP (Amount Percentage) = PayIn (Amount Percentage) * 0.8, rounded to up to 4 decimal
  places, with trailing zeros stripped. If PayIn Amount is 0, POSP Amount is also 0.
- POSP (OD Amount) and POSP (TP Amount) are always "0".
- Effect Start Date / Effect End Date are passed through unchanged from the batch context.

OUTPUT FORMAT — CRITICAL:
Return ONLY a JSON array, nothing else — no markdown fences, no commentary, no preamble.
Each element of the array is one output row, an object with EXACTLY these keys (all of them,
even if the value is null):
["rule_code","rule_name","ic_code","product_type","group","rule_type","cover_type",
"business_type","vehicle_age","state","rto","vehicle_category","vehicle_type","fuel_type",
"make","model","owner_type","usage_type","booking_mode","cover_selection_type","covers",
"addon_selection_type","addons","cc_from","cc_to","power_from","power_to","gvw_from","gvw_to",
"carrying_from","carrying_to","ncb_type","ncb_from","ncb_to","idv_from","idv_to",
"od_discount_from","od_discount_to","effect_start_date","effect_end_date",
"payin_commission_type","payin_reward_type","payin_amount_percentage","payin_od_amount",
"payin_tp_amount","posp_commission_type","posp_reward_type","posp_amount_percentage",
"posp_od_amount","posp_tp_amount"]

Use null (not the string "null") for blank cells. Use string values for everything else,
including numbers (e.g. "22.5" not 22.5), to avoid Excel formatting surprises.
"""

# ──────────────────────────────────────────────────────────────
# Groq client wrapper
# ──────────────────────────────────────────────────────────────

def get_groq_client():
    if Groq is None:
        raise RuntimeError(
            "groq package not installed. Run: pip install groq --break-system-packages"
        )
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise RuntimeError(
            "GROQ_API_KEY environment variable not set. Get a free key at "
            "https://console.groq.com/keys"
        )
    # CRITICAL: explicit timeout. Without one, a hung connection (network
    # hiccup, Groq being slow on a particular request, etc.) blocks forever
    # with no exception ever raised — call_llm_batch's try/except never
    # triggers, nothing logs, and the whole pipeline freezes on one batch
    # indefinitely. This is what "stuck at 0% forever, same log line
    # repeating, no error" in production was actually caused by. A 45s
    # timeout is generous for what's normally a few-second completion, but
    # short enough that a genuinely stuck call gets caught and retried
    # instead of stalling the whole multi-state run.
    return Groq(api_key=api_key, max_retries=0, timeout=45.0)


def _strip_json_fences(text):
    text = text.strip()
    text = re.sub(r"^```(json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    return text


def _repair_truncated_json_array(text):
    """Best-effort recovery for a JSON array that got cut off mid-object by
    max_tokens (the "Unterminated string..." / "Expecting property name..."
    errors seen in production). Rather than discarding an entire batch
    because the LAST row in it got cut off, walk backward from the parse
    error to the last complete '},' boundary, close the array there, and
    parse THAT. Returns a list of row dicts (possibly missing the last 1-2
    rows of the batch) or raises if nothing usable can be salvaged.

    This only ever throws away a row that was already incomplete/unusable —
    it never invents or alters data, it just stops being all-or-nothing
    about a single dropped row at the end of a batch.
    """
    last_obj_end = text.rfind("},")
    if last_obj_end == -1:
        # Maybe it's a single complete object with nothing after — try the
        # last '}' that isn't the very end (which already failed to parse).
        last_obj_end = text.rfind("}")
        if last_obj_end == -1:
            raise ValueError("no complete JSON object found to salvage")
        candidate = text[: last_obj_end + 1] + "]"
    else:
        candidate = text[: last_obj_end + 1] + "]"
    parsed = json.loads(candidate)
    if not isinstance(parsed, list) or not parsed:
        raise ValueError("salvaged JSON was not a usable non-empty array")
    return parsed


def call_llm_batch(client, sheet_name, batch_rows, rto_reference,
                    state_filter, eff_start, eff_end, naming_context):
    """
    Send one batch of raw grid rows to the LLM, return list of output row dicts.
    Retries on transient errors / malformed JSON, with backoff that's aware
    of *why* the call failed (rate limit vs. truncated output vs. other).
    """
    labeled_rows = label_batch_rows(sheet_name, batch_rows)

    user_payload = {
        "sheet": sheet_name,
        "state_filter": state_filter,
        "effect_start_date": eff_start,
        "effect_end_date": eff_end,
        "naming_instructions": naming_context,
        "rto_reference_for_this_state": rto_reference,
        "rows": labeled_rows,
    }

    user_msg = (
        f"Process these raw rows from sheet '{sheet_name}'. "
        f"Only generate rows relevant to state '{state_filter}'. "
        f"Effect Start Date = {eff_start}, Effect End Date = {eff_end}.\n\n"
        f"Naming pattern to follow for rule_name: {naming_context}\n\n"
        f"RTO reference for this state (cluster -> rto codes): "
        f"{json.dumps(rto_reference)}\n\n"
        f"Raw rows, each as a labeled object (JSON):\n{json.dumps(labeled_rows, default=str)}\n\n"
        f"Return the JSON array of output rows now."
    )

    multiplier = SHEET_OUTPUT_MULTIPLIER.get(sheet_name, 1)
    est_output_rows = max(len(batch_rows) * multiplier, 1)
    max_tokens = min(
        MAX_TOKENS_CEILING,
        max(MIN_MAX_TOKENS, est_output_rows * TOKENS_PER_OUTPUT_ROW),
    )

    # Rough total-token estimate for THIS call (system prompt + user payload +
    # expected completion). This used to only feed the client-side pacing
    # limiter, which paces BETWEEN calls but never shrinks a single call —
    # so a call whose own prompt+completion estimate exceeded the account's
    # TPM cap got waved through by the limiter and then hard-rejected by
    # Groq's server with a 413 ("Request too large... Limit 6000, Requested
    # 6247"), no retry possible since the call body itself is the problem.
    # Fix: shrink max_tokens here, BEFORE the call, so the request itself
    # can never exceed the cap on its own — same idea as the clamp already
    # in _RateLimiter.wait(), just applied to what we actually send instead
    # of just what we tell the pacer to expect.
    est_prompt_tokens = (len(RULES_PROMPT) + len(user_msg)) // 4  # ~4 chars/token
    tpm_safety_cap = int(GROQ_TPM_LIMIT * 0.9)  # headroom under the account's real cap
    available_for_completion = tpm_safety_cap - est_prompt_tokens
    if available_for_completion < MIN_MAX_TOKENS:
        # The prompt itself (system rules + this batch's RTO reference +
        # rows) is already eating most of the TPM budget — almost always
        # because this batch's cluster has an unusually large RTO list.
        # Forcing max_tokens up to MIN_MAX_TOKENS here would recreate the
        # exact 413 this fix exists to prevent, so instead we let
        # max_tokens go as low as the budget allows (down to a hard floor
        # of 150, enough for a single small row) rather than guarantee an
        # oversized request. A batch this constrained will likely still
        # need the JSON-truncation salvage path, but it'll get a real
        # response instead of an instant server-side rejection.
        max_tokens = max(150, available_for_completion)
        print(f"WARNING: prompt for {sheet_name}/{state_filter} batch is large "
              f"(~{est_prompt_tokens} tokens) — completion budget squeezed to "
              f"{max_tokens} tokens to stay under TPM cap")
    else:
        max_tokens = max(MIN_MAX_TOKENS, min(max_tokens, available_for_completion))
    est_call_tokens = est_prompt_tokens + max_tokens

    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        _rate_limiter.wait(estimated_tokens=est_call_tokens)
        try:
            resp = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": RULES_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
                temperature=0,
                max_tokens=max_tokens,
            )
            raw = resp.choices[0].message.content

            usage = getattr(resp, "usage", None)
            actual_tokens = getattr(usage, "total_tokens", None) if usage else None
            if actual_tokens:
                _rate_limiter.record_actual(est_call_tokens, actual_tokens)

            cleaned = _strip_json_fences(raw)
            print("\n========== RAW LLM RESPONSE ==========")
            print(cleaned)
            print("=====================================\n")
            try:
                parsed = json.loads(cleaned)
            except json.JSONDecodeError as parse_err:
                # Try to salvage the rows that DID complete before the
                # response got cut off, instead of throwing the whole batch
                # away on attempt 1. Only do this if the finish_reason
                # confirms we actually hit the token limit (so we don't
                # mask a genuine "the model wrote bad JSON" bug as if it
                # were a truncation).
                finish_reason = getattr(resp.choices[0], "finish_reason", None)
                if finish_reason == "length":
                    try:
                        salvaged = _repair_truncated_json_array(cleaned)
                        print(f"Recovered {len(salvaged)} row(s) from a "
                              f"truncated response (finish_reason=length)")
                        parsed = salvaged
                    except Exception:
                        raise parse_err
                else:
                    raise parse_err

            if not isinstance(parsed, list):
                raise ValueError("LLM did not return a JSON array")
            return parsed

        except _GroqRateLimitError as e:
            # Genuine 429 — back off for whatever Groq tells us (or a safe
            # default), and push the shared rate limiter out so the *next*
            # batch waits too, instead of immediately re-hitting the cap.
            retry_after = _extract_retry_after(e, default=RETRY_DELAY_SEC * attempt * 5)
            print(f"\n429 rate limited (attempt {attempt}/{MAX_RETRIES}); "
                  f"waiting {retry_after:.1f}s before retrying")
            last_err = e
            _rate_limiter.penalize(retry_after)
            time.sleep(retry_after)
            continue

        except json.JSONDecodeError as e:
            # Almost always a truncated response — give the next attempt
            # more room instead of just blindly retrying with the same budget.
            print(f"\nJSON parse error (attempt {attempt}/{MAX_RETRIES}): {e}")
            last_err = e
            max_tokens = min(MAX_TOKENS_CEILING, int(max_tokens * 1.6) + 200)
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY_SEC)
            continue

        except Exception as e:
            print("\nERROR:", e)
            last_err = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY_SEC * attempt)
            continue

    raise RuntimeError(
        f"LLM batch failed after {MAX_RETRIES} attempts for sheet={sheet_name}, "
        f"state={state_filter}: {last_err}"
    )


# ──────────────────────────────────────────────────────────────
# Source workbook loading (raw rows only — no field-mapping logic)
# ──────────────────────────────────────────────────────────────

RTO_STATE_NAMES = {
    "AN": "ANDAMAN ISLANDS", "AP": "ANDHRA PRADESH", "AR": "ARUNACHAL PRADESH",
    "AS": "ASSAM", "BR": "BIHAR", "CG": "CHHATTISGARH", "CH": "CHANDIGARH",
    "DD": "DAMAN AND DIU", "DL": "DELHI", "DN": "DADRA AND NAGAR HAVELI",
    "GA": "GOA", "GJ": "GUJARAT", "HP": "HIMACHAL PRADESH", "HR": "HARYANA",
    "JH": "JHARKHAND", "JK": "JAMMU KASHMIR", "KA": "KARNATAKA", "KL": "KERALA",
    "LA": "LADAKH", "LD": "LAKSHADWEEP", "MH": "MAHARASHTRA", "ML": "MEGHALAYA",
    "MN": "MANIPUR", "MP": "MADHYA PRADESH", "MZ": "MIZORAM", "NL": "NAGALAND",
    "OD": "ODISHA", "OR": "ODISHA", "PB": "PUNJAB", "PY": "PUDUCHERRY",
    "RJ": "RAJASTHAN", "SK": "SIKKIM", "TG": "TELANGANA", "TN": "TAMIL NADU",
    "TR": "TRIPURA", "TS": "TELANGANA", "UA": "UTTARAKHAND", "UK": "UTTARAKHAND",
    "UP": "UTTAR PRADESH", "WB": "WEST BENGAL",
}


def load_raw_sheets(path):
    """Load raw rows from each relevant sheet — no interpretation, just data."""
    wb = load_workbook(path, read_only=True, data_only=True)
    data = {}

    ws_rto = wb["2W RTO's"]
    rto_rows = list(ws_rto.iter_rows(values_only=True))
    data["rto_rows"] = [r for r in rto_rows[2:] if r and r[1]]

    for sheet_name, key in [
        ("TW 1+5", "tw_1p5"),
        ("TW 1+1 & SATP", "tw_1p1"),
        ("TW SAOD with Flexi Options", "tw_saod"),
    ]:
        ws = wb[sheet_name]
        data[key] = [list(r) for r in ws.iter_rows(values_only=True) if r and r[1]]

    wb.close()
    return data


def build_rto_lookup(rto_rows):
    """cluster -> {1p1: [rtos], 1p5: [rtos], saod: [rtos]}, and rto -> state."""
    lookup = {"1p1": {}, "1p5": {}, "saod": {}}
    rto_to_state = {}
    for r in rto_rows:
        rto = r[1]
        if not rto:
            continue
        state = RTO_STATE_NAMES.get(str(rto)[:2].upper())
        if state:
            rto_to_state[rto] = state
        for idx, prod in zip((2, 3, 4), ("1p1", "1p5", "saod")):
            cluster = r[idx] if len(r) > idx else None
            if cluster:
                key = str(cluster).strip().upper() if prod == "saod" else str(cluster).strip()
                lookup[prod].setdefault(key, []).append(rto)
    return lookup, rto_to_state


def get_all_states(rto_to_state):
    return sorted(set(rto_to_state.values()))


# Field names for each raw sheet's columns, by position, matching the
# layout load_raw_sheets() reads (index 0 is always the leading blank/merge
# column from the source workbook). Sending rows to the LLM as labeled
# objects instead of bare positional arrays removes a major source of
# inconsistent/incorrect extraction — the LLM no longer has to infer "is
# index 2 the Make or the Segment on THIS sheet" from prose alone, which
# differed between TW 1+5, TW 1+1 & SATP, and SAOD.
SHEET_FIELD_LABELS = {
    "TW 1+5": [
        None, "cluster", "make", "segment", "cd1", "max_cd2", "formula_type",
    ],
    "TW 1+1 & SATP": [
        None, "cluster", "segment", "cd1",
        "max_cd2_1plus1_od", "max_cd2_satp_tp",
    ],
    "TW SAOD with Flexi Options": [
        None, "cluster", "segment", "min_cd1",
        "max_cd1_no_breakin", "max_cd1_breakin", "cd2",
        "max_cd2_year1", "max_cd2_year2", "max_cd2_year3", "max_cd2_year4",
    ],
}


def label_batch_rows(sheet_name, batch_rows):
    """Convert raw positional rows into labeled dicts for this sheet, so the
    LLM reads named fields (e.g. "make", "segment", "max_cd2_year3") instead
    of having to infer column meaning from position."""
    labels = SHEET_FIELD_LABELS.get(sheet_name)
    if not labels:
        return batch_rows
    out = []
    for row in batch_rows:
        d = {}
        for i, val in enumerate(row):
            if i >= len(labels):
                # Trailing columns beyond what we know how to label for this
                # sheet — these are blank merge artifacts in every sheet
                # we've seen, so drop rather than burning tokens on
                # "col_11": null style noise.
                continue
            name = labels[i]
            if name is None:
                continue
            d[name] = val
        out.append(d)
    return out


def clusters_for_state(rto_rows_data, prod, state_name, rto_to_state):
    """Return {cluster: [rtos in this state]} for clusters touching this state."""
    out = {}
    for r in rto_rows_data:
        rto = r[1]
        if not rto or rto_to_state.get(rto) != state_name:
            continue
        idx = {"1p1": 2, "1p5": 3, "saod": 4}[prod]
        cluster = r[idx] if len(r) > idx else None
        if cluster:
            key = str(cluster).strip().upper() if prod == "saod" else str(cluster).strip()
            out.setdefault(key, []).append(rto)
    return out


# ──────────────────────────────────────────────────────────────
# Row dict -> Excel row list
# ──────────────────────────────────────────────────────────────

def row_dict_to_list(d):
    return [d.get(f) for f in JSON_FIELDS]


# ──────────────────────────────────────────────────────────────
# Output writer
# ──────────────────────────────────────────────────────────────

def write_output_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OUTPUT_HEADERS)
    hfill = PatternFill("solid", fgColor="366092")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 45)
    wb.save(path)


# ──────────────────────────────────────────────────────────────
# Main generation pipeline (per state)
# ──────────────────────────────────────────────────────────────

def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


_RULE_NAME_NUM_RE = re.compile(r"^(.*_)(\d+)(_TW)$")


def _renumber_rule_name(row_dict, counter):
    """Overwrite whatever numeric suffix the LLM produced with a sequential,
    state-wide counter (shared across every sheet/batch for this state).

    The LLM has no memory between batch calls — each call starts a fresh
    context, so "keep incrementing across batches" in the prompt text is
    not something it can actually do reliably. That's why batches were
    restarting at _1_TW each time. Counting in Python instead makes
    uniqueness/contiguity a guarantee rather than a hope.
    """
    name = row_dict.get("rule_name") or ""
    counter[0] += 1
    m = _RULE_NAME_NUM_RE.match(name)
    if m:
        row_dict["rule_name"] = f"{m.group(1)}{counter[0]}{m.group(3)}"
    else:
        # Unexpected pattern from the LLM — still keep numbering in sync
        # rather than silently producing another collision.
        row_dict["rule_name"] = f"{name or 'RULE'}_{counter[0]}_TW"
    return row_dict


def generate_for_state_ai(client, raw, rto_lookup, rto_to_state,
                           state_name, eff_start, eff_end,
                           progress_callback=None):
    all_rows = []
    failed_batches = []  # list of {"sheet": ..., "rows": n, "error": str}
    rule_counter = [0]  # mutable int shared across all sheets/batches below

    sheet_configs = [
        ("TW 1+5", "tw_1p5", "1p5",
         "rule_name pattern: '{CLUSTER_ABBR}_new_com_{N}_TW' where N increments per row "
         "you generate within this state's file (start from 1, keep incrementing across "
         "all sheets for this state)."),
        ("TW 1+1 & SATP", "tw_1p1", "1p1",
         "rule_name pattern: for the OD/renewal row use "
         "'{CLUSTER_ABBR}_all_RR_{N}_TW', for the TP row use "
         "'{CLUSTER_ABBR}_all_TP_{N}_TW' — N keeps incrementing across all rows/sheets "
         "for this state."),
        ("TW SAOD with Flexi Options", "tw_saod", "saod",
         "rule_name pattern: '{CLUSTER_ABBR}_all_Od_{N}_TW' — N keeps incrementing across "
         "all rows/sheets for this state. Generate 4 rows per segment (years 1-4)."),
    ]

    for sheet_name, data_key, prod_key, naming in sheet_configs:
        sheet_rows = raw[data_key]
        clusters = clusters_for_state(raw["rto_rows"], prod_key, state_name, rto_to_state)
        if not clusters:
            continue

        # Build cluster lookup for filtering sheet_rows to only this state's clusters
        cluster_keys = set(clusters.keys())

        def row_cluster_key(r):
            c = r[1]
            if not c:
                return None
            return str(c).strip().upper() if prod_key == "saod" else str(c).strip()

        relevant_rows = [r for r in sheet_rows if row_cluster_key(r) in cluster_keys]
        if not relevant_rows:
            continue

        batch_size = BATCH_SIZE_BY_SHEET.get(sheet_name, DEFAULT_BATCH_SIZE)
        for batch in chunk(relevant_rows, batch_size):
            if progress_callback:
                progress_callback(f"{state_name}: {sheet_name} batch...", 0, 0)
            # Only send RTO data for the cluster(s) actually present in THIS
            # batch, not the whole state's cluster table. A batch is usually
            # 1-2 rows, so this stays tiny regardless of how many RTOs the
            # state has overall (the prior `rto_ref = {}` was a blunt fix for
            # the 413/TPM error — it killed RTO accuracy as a side effect
            # instead of just shrinking the payload).
            batch_cluster_keys = {row_cluster_key(r) for r in batch if row_cluster_key(r)}
            rto_ref = {ck: clusters[ck] for ck in batch_cluster_keys if ck in clusters}
            try:
                results = call_llm_batch(
                    client, sheet_name, batch, rto_ref,
                    state_name, eff_start, eff_end, naming,
                )
            except RuntimeError as e:
                if progress_callback:
                    progress_callback(f"ERROR (batch dropped): {e}", 0, 0)
                failed_batches.append({
                    "sheet": sheet_name,
                    "rows": len(batch),
                    "error": str(e),
                })
                continue
            for r in results:
                _renumber_rule_name(r, rule_counter)
                all_rows.append(row_dict_to_list(r))

    return all_rows, failed_batches


def process_all_ai(input_path, output_dir, eff_start, eff_end,
                    states=None, progress_callback=None):
    os.makedirs(output_dir, exist_ok=True)
    client = get_groq_client()

    if progress_callback:
        progress_callback("Loading source workbook...", 0, 1)
    raw = load_raw_sheets(input_path)
    rto_lookup, rto_to_state = build_rto_lookup(raw["rto_rows"])
    all_states = get_all_states(rto_to_state)
    targets = [s for s in all_states if (states is None or s in states)]

    # Rough call-count estimate up front (helps catch a "this will take 2
    # hours" surprise before it actually takes 2 hours). Not exact — a row
    # can belong to more than one state for multi-state clusters like NCR —
    # but close enough to sanity-check before a long run.
    est_calls = 0
    for sheet_name, data_key, prod_key, _ in [
        ("TW 1+5", "tw_1p5", "1p5", None),
        ("TW 1+1 & SATP", "tw_1p1", "1p1", None),
        ("TW SAOD with Flexi Options", "tw_saod", "saod", None),
    ]:
        batch_size = BATCH_SIZE_BY_SHEET.get(sheet_name, DEFAULT_BATCH_SIZE)
        for state in targets:
            clusters = clusters_for_state(raw["rto_rows"], prod_key, state, rto_to_state)
            if not clusters:
                continue
            cluster_keys = set(clusters.keys())
            relevant = [
                r for r in raw[data_key]
                if r[1] and (str(r[1]).strip().upper() if prod_key == "saod" else str(r[1]).strip()) in cluster_keys
            ]
            if relevant:
                est_calls += -(-len(relevant) // batch_size)  # ceil division

    eta_min = (est_calls / GROQ_RPM_LIMIT) if GROQ_RPM_LIMIT else 0
    if progress_callback:
        progress_callback(
            f"Estimated ~{est_calls} Groq calls, ~{eta_min:.1f} min at {GROQ_RPM_LIMIT:.0f} RPM "
            f"(longer if retries/429s happen).", 0, len(targets),
        )

    generated = []
    all_failures = {}  # state -> list of failed batch dicts
    for idx, state in enumerate(targets):
        if progress_callback:
            progress_callback(f"Processing {state}...", idx, len(targets))
        rows, failed = generate_for_state_ai(
            client, raw, rto_lookup, rto_to_state, state,
            eff_start, eff_end, progress_callback,
        )
        if failed:
            all_failures[state] = failed
        if not rows:
            continue
        fname = state.replace(", ", "_").replace(" ", "") + "_2W.xlsx"
        out_path = os.path.join(output_dir, fname)
        write_output_excel(rows, out_path)
        generated.append(out_path)

    total_failed_batches = sum(len(v) for v in all_failures.values())
    if progress_callback:
        if total_failed_batches:
            progress_callback(
                f"Done with errors: {total_failed_batches} batch(es) across "
                f"{len(all_failures)} state(s) failed and were skipped — output is "
                f"INCOMPLETE for those states. See failure details.",
                len(targets), len(targets),
            )
        else:
            progress_callback("Complete!", len(targets), len(targets))

    return generated, all_failures


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)
    input_path, output_dir, eff_s, eff_e = sys.argv[1:5]
    states = sys.argv[5].split(",") if len(sys.argv) > 5 else None

    def cb(msg, cur, total):
        print(f"[{cur}/{total}] {msg}")

    files, failures = process_all_ai(input_path, output_dir, eff_s, eff_e, states, cb)
    print(f"\nGenerated {len(files)} files:")
    for f in files:
        print(f"  {f}")
    if failures:
        total_failed = sum(len(v) for v in failures.values())
        print(f"\n⚠ {total_failed} batch(es) FAILED and were skipped "
              f"(output is incomplete for these states):")
        for state, batches in failures.items():
            print(f"  {state}: {len(batches)} failed batch(es)")
            for b in batches:
                print(f"    - {b['sheet']} ({b['rows']} rows): {b['error']}")
        sys.exit(1)
