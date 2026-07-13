"""
Digit 2W Insurance Commission Grid Processor  v5
Key fixes:
  - POSP = PayIn * 0.8 always (confirmed from reference)
  - Fuel order: Petrol, LPG, Diesel, CNG  (matches reference)
  - Power values stored as formatted strings: '3.00', '7.00', etc.
  - SAOD Royal Enfield VehicleType = Bike (not ALL)
  - SAOD cluster case-insensitive matching fixed
  - Make field: 'Others' -> EXCLUDE list with reference brand order
    (HERO MOTOCORP, BAJAJ, HONDA, ROYAL ENFIELD, TVS, SUZUKI, YAMAHA —
    NOT alphabetical; fixed v4, was a "Make Issue" in QA review)
  - Rule name: STATE_new_com_N_TW format using cluster abbreviation
  - 1+1 ordering: all RR rows, then all TP rows per cluster block
  - EV bonus rows at end of each 1+1 cluster block
  - TW 1+5 generic EV power-band rows (< 3KW, 3-7KW, >7KW, plain EV) use
    Vehicle Type = Bike, NOT ALL. CONFIRMED v5: a v4 change that set these
    to ALL was WRONG and has been reverted. Reasoning (per business rule,
    confirmed by QA screenshot): the plain "SCOOTER" segment (Vehicle
    Type=Scooter, Fuel=ALL) already covers scooters + electric, so a
    generic ALL+Electric row would double-count that combination. Scoping
    the generic EV bands to Bike avoids the overlap. This only applies to
    the TW 1+5 sheet — 1+1 EV bonus rows correctly stay Bike regardless.
  - 1+1 "MC_180-350_Other than RE" / "..._Others" segment now excludes
    only ROYAL ENFIELD with Model=ALL, instead of incorrectly excluding
    BAJAJ/HONDA/JAWA too and restricting Model to "BAJAJ|AVENGER" (fixed
    v4, was a "Make model Exclude Issue" in QA review). CONFIRMED v5: the
    HONDA/JAWA/Avenger row's own Model stays "BAJAJ|AVENGER" (not the
    full Bajaj model list) — this was already correct and unchanged.

  v6 fixes (QA review against ODISHA reference output):
  - TW 1+5 generic EV power-band rows (< 3KW, 3-7KW, >7KW, plain EV) are
    Vehicle Type = ALL, NOT Bike. The v5 "confirmed" reasoning above (that
    Bike avoids double-counting with the plain SCOOTER segment) was WRONG
    per the reference output; reverted back to ALL for these rows and for
    the Bajaj 3-7KW gap-filler row in gen_1p5.
  - TW 1+5 "> 3 KW" and "> 7 KW" are different bands (3.10-100.00 vs
    7.10-100.00 respectively), not the same band. The old code collapsed
    both into 3.10-100.00.
  - TW 1+5 "3-7 KW" no longer synthesizes compensating 0%-payin rows for
    the bands below 3KW / above 7KW. When a cluster/make has its own real
    "< 3 KW"/"< 7 KW" and "> 3 KW"/"> 7 KW" entries (e.g. NE_OR_Good), the
    old compensating rows duplicated and stomped those real rates with a
    bogus 0%.
  - 1+1 "SC/EV" segment Fuel Type is ALL, not Petrol-only — it must cover
    both petrol and electric scooters.
  - 1+1 "MC_180-350_Other than RE" / "..._Others" now excludes ROYAL
    ENFIELD, HONDA, and JAWA MOTORCYCLE (not just RE), and its Model is
    "EXCLUDE: BAJAJ|AVENGER" (not ALL), since HONDA/JAWA/BAJAJ|Avenger are
    already covered by the dedicated MC_180-350_HONDA/JAWA/Avenger row.
"""

import os
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

OUTPUT_HEADERS = [
    "Rule Code", "Rule Name *", "IC Code *", "Product Type *", "Group",
    "Rule Type *", "Cover Type", "Business Type", "Vehicle Age", "State",
    "RTO", "Vehicle Category", "Vehicle Type", "Fuel Type", "Make", "Model",
    "Owner Type", "Usage Type", "Booking Mode", "Cover Selection Type",
    "Covers", "Addon Selection Type", "Addons", "CC From", "CC To",
    "Power From", "Power To", "GVW From", "GVW To", "Carrying From",
    "Carrying To", "NCB Type", "NCB From", "NCB To", "IDV From", "IDV To",
    "OD Discount From", "OD Discount To", "Effect Start Date *",
    "Effect End Date *", "PayIn (Commision Type)", "PayIn (Reward Type)",
    "PayIn (Amount Percentage)", "PayIn (OD Amount)", "PayIn (TP Amount)",
    "POSP (Commision Type)", "POSP (Reward Type)", "POSP (Amount Percentage)",
    "POSP (OD Amount)", "POSP (TP Amount)"
]

MISP_RATE = 22.5
FUEL_PETROL = "Petrol, CNG, LPG, Diesel"   # Fixed order to match reference output
FUEL_ALL    = "ALL"
FUEL_EV     = "Electric"
EXCLUDE_OTHERS = "EXCLUDE: HERO MOTOCORP, BAJAJ, HONDA, ROYAL ENFIELD, TVS, SUZUKI, YAMAHA"  # reference brand order (not alphabetical)

# Canonical brand list used for computing EXCLUDE when "Others" appears in slash groups
# Order matters: preserve alphabetical for EXCLUDE computations
KNOWN_BRANDS_ORDERED = ["BAJAJ", "HERO MOTOCORP", "HONDA", "JAWA MOTORCYCLE",
                        "ROYAL ENFIELD", "SUZUKI", "TVS", "YAMAHA"]
KNOWN_BRANDS_SET = {b.upper() for b in KNOWN_BRANDS_ORDERED}

# Full model string for the Avenger row (MC_180-350_HONDA/JAWA/Avenger).
# NOTE: no longer used in make_field_1p1 as of the latest fix — that row's
# Model is now just "BAJAJ|AVENGER" (see change 3). Kept here for reference
# only; safe to delete if it's not needed elsewhere.
AVENGER_EXCLUDE_MODEL = (
    "EXCLUDE: BAJAJ|ASPIRE, BAJAJ|BOXER, BAJAJ|BOXER 150, BAJAJ|BYK, BAJAJ|CALIBER, "
    "BAJAJ|CT 100, BAJAJ|CUB, BAJAJ|DISCOVER, BAJAJ|DISCOVER 112, BAJAJ|DISCOVER 125, "
    "BAJAJ|DISCOVER 135 DTSI, BAJAJ|DISCOVER 150, BAJAJ|DISCOVER F, BAJAJ|DISCOVER M, "
    "BAJAJ|DISCOVER S, BAJAJ|DISCOVER ST, BAJAJ|DISCOVER T, BAJAJ|ELIMINATOR, "
    "BAJAJ|KB 100, BAJAJ|KB 125 RTZ, BAJAJ|PLATINA, BAJAJ|PULSAR, BAJAJ|SONIC, "
    "BAJAJ|SX ENDURO, BAJAJ|V15, BAJAJ|4S CHAMPION, BAJAJ|WIND 125, BAJAJ|XCD, "
    "BAJAJ|XCD 125, BAJAJ|XCD 135, BAJAJ|XCD 135 DTSI, BAJAJ|V12, BAJAJ|DOMINAR, "
    "BAJAJ|CT 110, BAJAJ|DISCOVER 110, BAJAJ|CT 125, LML|FREEDOM, BAJAJ|BRAVO, "
    "BAJAJ|CHETAK, ROYAL ENFIELD|CLASSIC, BAJAJ|M 80, BAJAJ|PRIYA, BAJAJ|SAFFIRE, "
    "BAJAJ|SPIRIT, BAJAJ|SUNNY, BAJAJ|SUPER, BAJAJ|BLADE, BAJAJ|WAVE, BAJAJ|KRISTAL, "
    "BAJAJ|LEGEND, YAMAHA|NEO"
)
AVENGER_MODEL = (
    "BAJAJ|ASPIRE, BAJAJ|BOXER, BAJAJ|BOXER 150, BAJAJ|BYK, BAJAJ|CALIBER, "
    "BAJAJ|CT 100, BAJAJ|CUB, BAJAJ|DISCOVER, BAJAJ|DISCOVER 112, BAJAJ|DISCOVER 125, "
    "BAJAJ|DISCOVER 135 DTSI, BAJAJ|DISCOVER 150, BAJAJ|DISCOVER F, BAJAJ|DISCOVER M, "
    "BAJAJ|DISCOVER S, BAJAJ|DISCOVER ST, BAJAJ|DISCOVER T, BAJAJ|ELIMINATOR, "
    "BAJAJ|KB 100, BAJAJ|KB 125 RTZ, BAJAJ|PLATINA, BAJAJ|PULSAR, BAJAJ|SONIC, "
    "BAJAJ|SX ENDURO, BAJAJ|V15, BAJAJ|4S CHAMPION, BAJAJ|WIND 125, BAJAJ|XCD, "
    "BAJAJ|XCD 125, BAJAJ|XCD 135, BAJAJ|XCD 135 DTSI, BAJAJ|V12, BAJAJ|DOMINAR, "
    "BAJAJ|CT 110, BAJAJ|DISCOVER 110, BAJAJ|CT 125, LML|FREEDOM, ROYAL ENFIELD|CLASSIC, "
    "BAJAJ|BRAVO, BAJAJ|CHETAK, BAJAJ|M 80, BAJAJ|PRIYA, BAJAJ|SAFFIRE, "
    "BAJAJ|SPIRIT, BAJAJ|SUNNY, BAJAJ|SUPER, BAJAJ|BLADE, BAJAJ|WAVE, BAJAJ|KRISTAL, "
    "BAJAJ|LEGEND, YAMAHA|NEO"
)


def normalize_make_1p5(make_raw):
    """
    Transform TW 1+5 Make cell → correct output string.

    Rules (from reference Digit_2w.zip):
    1. Plain 'Others'  → EXCLUDE_OTHERS (all 7 brands, alphabetical)
    2. Slash group WITHOUT Others (e.g. 'BAJAJ/HONDA') → 'BAJAJ, HONDA' (comma-space)
    3. Slash group WITH Others (e.g. 'TVS/SUZUKI/Others') →
       'EXCLUDE: <brands NOT in the named list>, in KNOWN_BRANDS_ORDERED order'
       i.e. the listed brands are the ones covered; everything else is excluded
    """
    raw = str(make_raw).strip()
    parts = [p.strip() for p in raw.split("/")]
    upper_parts = [p.upper() for p in parts]
    has_others = any(p == "OTHERS" for p in upper_parts)
    named_upper = [p for p in upper_parts if p != "OTHERS"]

    if not has_others:
        # Simple slash list → comma join (preserve original casing)
        non_others = [p for p in parts if p.upper() != "OTHERS"]
        return ", ".join(non_others)

    if not named_upper:
        # Pure 'Others' alone
        return EXCLUDE_OTHERS

    # Mixed: e.g. 'TVS/SUZUKI/Others/YAMAHA'
    # These named brands ARE covered by this row; the EXCLUDE is the rest
    excluded = [b for b in KNOWN_BRANDS_ORDERED
                if b.upper() not in named_upper and b != "JAWA MOTORCYCLE"]
    if excluded:
        return "EXCLUDE: " + ", ".join(excluded)
    return EXCLUDE_OTHERS

def _bajaj_only_brand(make_raw):
    """
    True if BAJAJ is the only *named* brand in make_raw once 'Others' is
    stripped out (e.g. 'Bajaj', 'BAJAJ', 'Bajaj/Others', 'Bajaj, Others'
    all → True).

    Some clusters (e.g. NE_OR_GOOD) record the Bajaj EV rows with a make
    cell like 'Bajaj/Others' instead of a plain 'Bajaj'. Left alone,
    normalize_make_1p5() reads that as an EXCLUDE list, which is wrong for
    a row that's fundamentally a Bajaj row. This lets callers force the
    make back to plain 'BAJAJ' for these rows.

    Splits on both '/' and ',' so minor source-formatting differences
    (slash-separated vs comma-separated) don't cause this to miss.
    """
    raw = str(make_raw).strip().replace(",", "/")
    parts = [p.strip().upper() for p in raw.split("/") if p.strip()]
    named = [p for p in parts if p != "OTHERS"]
    return len(named) == 1 and named[0] == "BAJAJ"


def _norm_seg(seg):
    """Collapse whitespace and upper-case a segment label for robust
    comparisons (so '< 3 KW', '<3 KW', '<  3 kw' all match the same
    way)."""
    return " ".join(str(seg).strip().split()).upper()

RTO_STATE_NAMES = {
    "AN": "ANDAMAN ISLANDS",   "AP": "ANDHRA PRADESH",
    "AR": "ARUNACHAL PRADESH", "AS": "ASSAM",
    "BR": "BIHAR",             "CG": "CHHATTISGARH",
    "CH": "CHANDIGARH",        "DD": "DAMAN AND DIU",
    "DL": "DELHI",             "DN": "DADRA AND NAGAR HAVELI",
    "GA": "GOA",               "GJ": "GUJARAT",
    "HP": "HIMACHAL PRADESH",  "HR": "HARYANA",
    "JH": "JHARKHAND",         "JK": "JAMMU KASHMIR",
    "KA": "KARNATAKA",         "KL": "KERALA",
    "LA": "LADAKH",            "LD": "LAKSHADWEEP",
    "MH": "MAHARASHTRA",       "ML": "MEGHALAYA",
    "MN": "MANIPUR",           "MP": "MADHYA PRADESH",
    "MZ": "MIZORAM",           "NL": "NAGALAND",
    "OD": "ODISHA",            "OR": "ODISHA",
    "PB": "PUNJAB",            "PY": "PUDUCHERRY",
    "RJ": "RAJASTHAN",         "SK": "SIKKIM",
    "TG": "TELANGANA",         "TN": "TAMIL NADU",
    "TR": "TRIPURA",           "TS": "TELANGANA",
    "UA": "UTTARAKHAND",       "UK": "UTTARAKHAND",
    "UP": "UTTAR PRADESH",     "WB": "WEST BENGAL",
}

CLUSTER_ABBR = {
    "Andaman": "ANDAMAN", "APTS_Bad": "TS_AP_Bad", "APTS_Good": "TS_AP_Good",
    "APTS_Good1": "TS_AP_Good1", "APTS_Good2": "TS_AP_Good2",
    "APTS_Ref": "TS_AP_ref", "BR_Bad": "BR", "BR_Good": "BR",
    "CG+MP Bad": "CG_MP", "CG,MP,HR,HP ,WB REF": "REF", "CG_Good": "CG",
    "Chandigarh": "CH", "DL_Delhi": "DL", "GJ_ABS": "GJ", "GJ_Bad": "GJ",
    "GJ_Good": "GJ", "Goa": "GA", "HP_Bad": "HP", "HP_Good": "HP",
    "HR_Bad": "HR", "HR_Good": "HR", "J&K_Bad": "JK", "J&K_Jammu": "JK",
    "J&K_Srinagar": "JK", "Jharkhand": "JH", "KA_Bad": "KA",
    "KA_Bangalore": "KA", "MH_Bad": "MH", "MH_Good": "MH",
    "MH_Mumbai": "MH", "MH_Pune": "MH", "MP_Good": "MP", "NCR": "NCR",
    "NE_Bad": "NE", "NE_OR_Good": "NE_OR", "NE_OR_GOOD": "NE_OR",
    "NE_Ref": "NE", "OR_Bad": "OR", "PB_Bad": "PB", "PB_Good": "PB",
    "RJ_Bad": "RJ", "RJ_Good": "RJ", "RJ_Jaipur": "RJ",
    "ROM+RJ+GJ+UP+UK+OR REF": "ROM_REF", "TN+KL Good": "TN_KL",
    "TN+KL+KA REF": "TN_KL_KA", "TN+KL_Bad": "TN_KL", "TN_Chennai": "TN",
    "UP+UK_Bad": "UP_UK", "UP+UK_Bad2": "UP_UK", "UP+UK_Good": "UP_UK",
    "UP+UK_Good2": "UP_UK", "WB_Bad": "WB", "WB_Good": "WB",
    "WB_Kolkata": "WB",
    # SAOD clusters
    "AHMEDABAD,SURAT,BARODA": "GJ", "AP": "AP", "Assam": "AS", "BR": "BR",
    "CG": "CG", "Chennai_Coimbatore": "TN", "DL": "DL", "GA": "GA",
    "GJ": "GJ", "HR": "HR", "HP": "HP", "JK": "JK", "JH": "JH",
    "KA": "KA", "KL": "KL", "MP": "MP", "MH_Bad": "MH", "MH": "MH",
    "MH_Mumbai,PUNE": "MH", "NE": "NE", "OR": "OR", "PB": "PB",
    "PY": "PY", "RJ": "RJ", "TN": "TN", "TG": "TG", "UP": "UP",
    "UK": "UK", "WB": "WB",
}

CLUSTER_ABBR_CI = {k.strip().upper(): v for k, v in CLUSTER_ABBR.items()}

def _abbr(cluster):
    hit = CLUSTER_ABBR_CI.get(str(cluster).strip().upper())
    if hit is not None:
        return hit
    return cluster.replace(" ", "").replace(",", "_").replace("+", "_")[:10].upper()

# ── PayIn helpers ─────────────────────────────────────────────

def _pct(val):
    if val is None or val == 0:
        return "0"
    if isinstance(val, str):
        return "0"
    v = round(float(val) * 100, 4)
    s = f"{v:.4f}".rstrip("0").rstrip(".")
    return s

def resolve_payin(cd2):
    if cd2 is None or cd2 == 0:
        return ("net", "0")
    if isinstance(cd2, str):
        u = cd2.strip().upper()
        if u == "MISP":
            return ("od", str(MISP_RATE))
        if u == "D":
            return ("net", "0")
        try:
            return ("net", _pct(float(cd2)))
        except Exception:
            return ("net", "0")
    return ("net", _pct(cd2))

def posp_amt(payin_pct):
    """POSP = PayIn * 0.8 always"""
    try:
        p = float(payin_pct)
    except Exception:
        return "0"
    if p == 0:
        return "0"
    r = round(p * 0.8, 4)
    return f"{r:.4f}".rstrip("0").rstrip(".")

def fmt_power(val):
    """Format power as '3.00' style string"""
    if val is None:
        return None
    return f"{float(val):.2f}"

# ── Row builder ───────────────────────────────────────────────

def build_row(rule_name, state, rto,
              cover_type, biz_type, veh_age,
              vtype, fuel, make, model,
              ccf, cct, pwf, pwt,
              payin_type, payin_pct,
              eff_start, eff_end):
    p = posp_amt(payin_pct)
    return [
        None, rule_name, "DIGIT", "TW", None,
        "PAYINPAYOUT", cover_type, biz_type, veh_age, state, rto,
        None, vtype, fuel, make, model,
        "ALL", None, "any", "na", None, "na", None,
        ccf, cct, pwf, pwt,
        None, None, None, None,
        "na", None, None, None, None, None, None,
        eff_start, eff_end,
        payin_type, "percentage", payin_pct, "0", "0",
        "net", "percentage", p, "0", "0",
    ]

# ── Data loading ──────────────────────────────────────────────

def load_input_data(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    d = {}

    ws_rto = wb["2W RTO's"]
    rto_rows = list(ws_rto.iter_rows(values_only=True))
    d["rto"] = {}
    for r in rto_rows[2:]:
        if r[1]:
            d["rto"][r[1]] = {"1p1": r[2], "1p5": r[3], "saod": r[4]}

    # Build cluster->rto lookups. Source sheets are inconsistent about
    # cluster-name casing (e.g. 'NE_OR_GOOD' for some brand rows vs
    # 'NE_OR_Good' for others, with only one spelling present in the RTO
    # sheet). An exact-string match would silently drop the mismatched
    # group's rows entirely (rtos == [] -> generator skips it). To avoid
    # that, ALL three product types key their cluster->rto lookup on the
    # case-normalized (stripped+uppercased) cluster name, not just saod.
    for prod in ("1p1", "1p5", "saod"):
        key = f"c_{prod}"
        d[key] = {}
        for rto, c in d["rto"].items():
            cl = c.get(prod)
            if cl:
                cl_key = cl.strip().upper()
                d[key].setdefault(cl_key, []).append(rto)

    ws = wb["TW 1+5"]
    d["tw_1p5"] = [
        {"cluster": str(r[1]).strip().upper(), "make": r[2], "seg": r[3], "cd2": r[5]}
        for r in ws.iter_rows(values_only=True) if r[1] and r[2] and r[3]
    ]

    ws = wb["TW 1+1 & SATP"]
    d["tw_1p1"] = [
        {"cluster": str(r[1]).strip().upper(), "seg": r[2], "cd2_1p1": r[4], "cd2_satp": r[5]}
        for r in ws.iter_rows(values_only=True) if r[1] and r[2] and r[3] is not None
    ]

    ws = wb["TW SAOD with Flexi Options"]
    d["tw_saod"] = [
        {"cluster": r[1], "seg": r[2],
         "yr1": r[7], "yr2": r[8], "yr3": r[9], "yr4": r[10]}
        for r in ws.iter_rows(values_only=True) if r[1] and r[2] and r[6] is not None
    ]

    wb.close()
    return d

def get_all_states(d):
    states = set()
    for rto in d["rto"]:
        s = RTO_STATE_NAMES.get(rto[:2].upper())
        if s:
            states.add(s)
    return sorted(states)

def _state_from_rtos(rtos):
    states = set()
    for rto in rtos:
        s = RTO_STATE_NAMES.get(rto[:2].upper())
        if s:
            states.add(s)
    return ", ".join(sorted(states))

def _rto_field(rtos):
    return ", ".join(sorted(rtos)) if rtos else "ALL"

# ── TW 1+5 segment → rows ─────────────────────────────────────

def seg_rows_1p5(make_raw, seg, cd2, rule_pfx, state, rto, counter, eff_s, eff_e):
    rows = []
    s = str(seg).strip()
    s_norm = _norm_seg(s)
    pt, pp = resolve_payin(cd2)
    make = normalize_make_1p5(make_raw)
    if _bajaj_only_brand(make_raw):
        # Force plain 'BAJAJ' even if the source cell was e.g. 'Bajaj/Others'
        make = "BAJAJ"

    def add(vt, fuel, ccf=None, cct=None, pwf=None, pwt=None, use_pp=True):
        nonlocal counter
        name = f"{rule_pfx}_new_com_{counter}_TW"
        rows.append(build_row(name, state, rto,
                              "Comprehensive", "new", "ALL",
                              vt, fuel, make, "ALL",
                              ccf, cct,
                              fmt_power(pwf), fmt_power(pwt),
                              pt if use_pp else "net",
                              pp if use_pp else "0",
                              eff_s, eff_e))
        counter += 1

    # EV power bands. Vehicle Type is ALL specifically for BAJAJ-only EV
    # power-band rows; every other make (including "Others"/EXCLUDE-brand
    # rows) stays Bike. CORRECTED (2nd pass): the first fix made ALL EV
    # power bands VehicleType=ALL regardless of make, but the reference
    # output shows the "Others" EXCLUDE-brand EV rows (e.g. NE_OR_Good)
    # are Bike, not ALL — only the BAJAJ EV rows are ALL. Only the
    # explicitly scooter-scoped "SCOOTER/EV" segment stays Scooter
    # regardless of make.
    #
    # Also fixed: "> 3 KW" and "> 7 KW" are NOT the same band. "< 3 KW"/
    # "< 7 KW" both represent the low band (0.10-2.90) regardless of
    # label, but the high-side label tells you whether this cluster's
    # split is at 3KW or 7KW, so ">3 KW" must map to 3.10-100.00 while
    # ">7 KW" maps to 7.10-100.00. The old code collapsed both into
    # 3.10-100.00, which produced a wrong/overlapping power range whenever
    # a cluster used the 7KW split together with an explicit "3-7 KW" mid
    # band (e.g. NE_OR_Good).
    #
    # Also removed here: the old "3-7 KW" branch used to also synthesize
    # compensating 0%-payin rows for the bands below 3KW and above 7KW
    # unconditionally. That's wrong whenever this cluster/make already has
    # its own real "< 3 KW"/"< 7 KW" and "> 3 KW"/"> 7 KW" entries (as
    # NE_OR_Good's "Others" make does) — it created duplicate, overlapping
    # rows that stomped the real rate with a bogus 0%. The mid band here
    # only ever comes from its own explicit "3-7 KW" entry; low/high bands
    # come from their own entries when present. Gap-filling for makes that
    # are genuinely missing a low or high entry (e.g. BAJAJ when only a
    # mid entry exists) is handled at the cluster level in gen_1p5(), which
    # has visibility across all of a make's entries in the cluster.
    ev_vt = "ALL" if _bajaj_only_brand(make_raw) else "Bike"
    if s_norm in ("< 3 KW", "< 7 KW"):
        add(ev_vt, FUEL_EV, pwf=0.10, pwt=2.90)
    elif s_norm in (">3 KW", "> 3 KW"):
        add(ev_vt, FUEL_EV, pwf=3.10, pwt=100.00)
    elif s_norm in (">7 KW", "> 7 KW"):
        add(ev_vt, FUEL_EV, pwf=7.10, pwt=100.00)
    elif s_norm == "3-7 KW":
        add(ev_vt, FUEL_EV, pwf=3.00, pwt=7.00)
    elif s_norm == "EV":
        add(ev_vt, FUEL_EV)
    elif s_norm == "SCOOTER/EV":
        add("Scooter", FUEL_EV)
    # Bike CC bands
    elif s_norm == "MC <=155":
        add("Bike", FUEL_PETROL, ccf=1, cct=155)
    elif s_norm == "MC >155":
        add("Bike", FUEL_PETROL, ccf=156, cct=9999)
    elif s_norm == "MC":
        add("Bike", FUEL_PETROL)
    elif s_norm == "SCOOTER":
        add("Scooter", FUEL_ALL)
    elif s_norm == "SCOOTER/MC":
        add("ALL", FUEL_PETROL)
    elif s_norm in ("< =350", "<= 350", "<=350"):
        add("ALL", FUEL_PETROL, ccf=1, cct=350)
    elif s_norm in ("> 350", ">350"):
        add("ALL", FUEL_PETROL, ccf=351, cct=9999)
    elif s_norm == "ALL":
        add("ALL", FUEL_ALL)
    else:
        add("ALL", FUEL_ALL)

    return rows, counter

# ── 1+1 segment info ──────────────────────────────────────────

def veh_info_1p1(seg):
    s = str(seg).strip()
    if s == "SC/EV":
        # "SC/EV" = scooter, any fuel (covers both petrol and electric
        # scooters). CORRECTED: was scoped to Petrol-only, which silently
        # excluded electric scooters from this rule.
        return ("Scooter", FUEL_ALL, None, None)
    if s in ("MC <= 180 Hero/Honda", "MC <= 180 Hero/Honda/TVS", "MC <= 180 Others"):
        return ("Bike", FUEL_PETROL, 1, 180)
    if s in ("MC_180-350_RE", "MC_180-350_HONDA/JAWA/Avenger",
             "MC_180-350_Other than RE", "MC_180-350_Others"):
        return ("Bike", FUEL_PETROL, 181, 350)
    if s == "MC>350":
        return ("Bike", FUEL_PETROL, 351, 9999)
    return ("ALL", FUEL_ALL, None, None)

def make_field_1p1(seg):
    """Returns (make, model) tuple for TW 1+1 rows."""
    s = str(seg).strip()
    if s == "SC/EV":
        return "ALL", "ALL"
    if s == "MC <= 180 Hero/Honda":
        return "HERO MOTOCORP, HONDA", "ALL"
    if s == "MC <= 180 Hero/Honda/TVS":
        return "HERO MOTOCORP, HONDA, TVS", "ALL"
    if s == "MC <= 180 Others":
        return "EXCLUDE: HERO MOTOCORP, HONDA", "ALL"
    if s == "MC_180-350_RE":
        return "ROYAL ENFIELD", "ALL"
    if s == "MC_180-350_HONDA/JAWA/Avenger":
        # make = HONDA, JAWA MOTORCYCLE, BAJAJ; model is just the single
        # "BAJAJ|AVENGER" entry, not the long list of individual Bajaj
        # models.
        return "HONDA, JAWA MOTORCYCLE, BAJAJ", "BAJAJ|AVENGER"
    if s in ("MC_180-350_Other than RE", "MC_180-350_Others"):
        # "Other than RE" is the broad catch-all for this CC band. It must
        # exclude ROYAL ENFIELD, HONDA, and JAWA MOTORCYCLE — all three
        # have their own dedicated rows above (RE gets its own row; HONDA/
        # JAWA/BAJAJ|Avenger gets its own row) — and, for the makes it does
        # cover (chiefly BAJAJ), it must exclude the BAJAJ|AVENGER model
        # specifically, since that model is already covered by the
        # HONDA/JAWA/Avenger row. CORRECTED: a prior revision excluded only
        # Royal Enfield and left Model=ALL, which double-covered
        # HONDA/JAWA/Avenger with conflicting rates.
        return "EXCLUDE: ROYAL ENFIELD, HONDA, JAWA MOTORCYCLE", "EXCLUDE: BAJAJ|AVENGER"
    if s == "MC>350":
        return "ALL", "ALL"
    return "ALL", "ALL"

# ── SAOD segment ──────────────────────────────────────────────

def veh_info_saod(seg):
    s = str(seg).strip()
    if s == "MC <155":
        return ("Bike",   FUEL_PETROL, 1,    155,  None, None, "EXCLUDE: ROYAL ENFIELD")
    if s == "MC>155":
        return ("Bike",   FUEL_PETROL, 156,  9999, None, None, "EXCLUDE: ROYAL ENFIELD")
    if s == "RE":
        return ("Bike",   FUEL_ALL,    None, None, None, None, "ROYAL ENFIELD")
    if s == "SC":
        return ("Scooter",FUEL_PETROL, None, None, None, None, "EXCLUDE: ROYAL ENFIELD")
    if s == "SC_EV":
        return ("Scooter",FUEL_EV,     None, None, None, None, "EXCLUDE: ROYAL ENFIELD")
    return ("ALL", FUEL_ALL, None, None, None, None, "ALL")

# ── Generators ────────────────────────────────────────────────

def gen_1p5(d, eff_s, eff_e, counter, state_filter=None):
    rows = []
    clusters = {}
    for e in d["tw_1p5"]:
        clusters.setdefault(e["cluster"], []).append(e)

    for cluster in sorted(clusters):
        entries = clusters[cluster]
        rtos = sorted(d["c_1p5"].get(cluster, []))
        if not rtos:
            continue
        state = _state_from_rtos(rtos)
        if state_filter and state_filter not in state.split(", "):
            continue
        rto_f = _rto_field(rtos)
        pfx = _abbr(cluster)

        # EV power-band gap filling, per make within this cluster. Two
        # distinct gaps can occur in the source data:
        #   (a) low ("< 3 KW"/"< 7 KW") and high ("> 3 KW"/"> 7 KW") exist
        #       as separate rows but no combined "3-7 KW" mid row is given
        #       — synthesize the mid band using the low band's rate.
        #   (b) a "3-7 KW" mid row exists but no separate low and/or high
        #       row exists for that make — synthesize the missing side(s)
        #       at 0% payin (confirmed against reference output for BAJAJ).
        # This used to be hardcoded to BAJAJ only; generalized here since
        # the same gap shapes can occur for any make. A make's entries are
        # grouped by its *output* make text (BAJAJ forced for Bajaj-only
        # cells; normalize_make_1p5() otherwise) so that source-formatting
        # differences like "Bajaj/Others" vs "Bajaj" still group together.
        ev_groups = {}
        for e in entries:
            seg_s = _norm_seg(e["seg"])
            if seg_s not in ("< 3 KW", "< 7 KW", "3-7 KW",
                              ">3 KW", "> 3 KW", ">7 KW", "> 7 KW"):
                continue
            is_bajaj = _bajaj_only_brand(e["make"])
            mk = "BAJAJ" if is_bajaj else normalize_make_1p5(e["make"])
            grp = ev_groups.setdefault(mk, {"low": None, "mid": None,
                                             "high": None, "is_bajaj": is_bajaj})
            if seg_s in ("< 3 KW", "< 7 KW"):
                grp["low"] = e
            elif seg_s == "3-7 KW":
                grp["mid"] = e
            else:
                grp["high"] = e

        for e in entries:
            make_raw = e["make"]
            new_rows, counter = seg_rows_1p5(make_raw, e["seg"], e["cd2"],
                                             pfx, state, rto_f,
                                             counter, eff_s, eff_e)
            rows.extend(new_rows)

            # Gap (b): right after emitting a mid ("3-7 KW") row, fill in
            # any missing low/high sides for that make at 0% payin.
            if _norm_seg(e["seg"]) == "3-7 KW":
                is_bajaj = _bajaj_only_brand(make_raw)
                mk = "BAJAJ" if is_bajaj else normalize_make_1p5(make_raw)
                grp = ev_groups.get(mk, {})
                vt = "ALL" if is_bajaj else "Bike"
                make_final = "BAJAJ" if is_bajaj else normalize_make_1p5(make_raw)
                if grp.get("low") is None:
                    name = f"{pfx}_new_com_{counter}_TW"
                    rows.append(build_row(name, state, rto_f,
                                          "Comprehensive", "new", "ALL",
                                          vt, FUEL_EV, make_final, "ALL",
                                          None, None,
                                          fmt_power(0.10), fmt_power(2.90),
                                          "net", "0", eff_s, eff_e))
                    counter += 1
                if grp.get("high") is None:
                    name = f"{pfx}_new_com_{counter}_TW"
                    rows.append(build_row(name, state, rto_f,
                                          "Comprehensive", "new", "ALL",
                                          vt, FUEL_EV, make_final, "ALL",
                                          None, None,
                                          fmt_power(7.10), fmt_power(100.00),
                                          "net", "0", eff_s, eff_e))
                    counter += 1

        # Gap (a): makes with low+high but no mid entry at all.
        for mk, grp in ev_groups.items():
            if grp["mid"] is not None or grp["low"] is None or grp["high"] is None:
                continue
            # NOTE/assumption: the middle band's PayIn is taken from the
            # low-band entry's rate, since the source doesn't give a rate
            # for a combined 3-7 KW row in this case. Check this against
            # the reference output and adjust if the real rate should
            # differ for non-Bajaj makes.
            vt = "ALL" if grp["is_bajaj"] else "Bike"
            make_final = "BAJAJ" if grp["is_bajaj"] else normalize_make_1p5(grp["low"]["make"])
            pt, pp = resolve_payin(grp["low"]["cd2"])
            name = f"{pfx}_new_com_{counter}_TW"
            rows.append(build_row(name, state, rto_f,
                                  "Comprehensive", "new", "ALL",
                                  vt, FUEL_EV, make_final, "ALL",
                                  None, None,
                                  fmt_power(3.00), fmt_power(7.00),
                                  pt, pp, eff_s, eff_e))
            counter += 1

    return rows, counter


def gen_1p1(d, eff_s, eff_e, counter, state_filter=None):
    rows = []
    clusters = {}
    for e in d["tw_1p1"]:
        clusters.setdefault(e["cluster"], []).append(e)

    for cluster in sorted(clusters):
        entries = clusters[cluster]
        rtos = sorted(d["c_1p1"].get(cluster, []))
        if not rtos:
            continue
        state = _state_from_rtos(rtos)
        if state_filter and state_filter not in state.split(", "):
            continue
        rto_f = _rto_field(rtos)
        pfx = _abbr(cluster)

        rr_rows = []
        tp_rows = []
        has_ev = False
        ev_cd2_rr = None
        ev_cd2_tp = None

        for e in entries:
            seg = e["seg"]
            vt, fuel, ccf, cct = veh_info_1p1(seg)
            mf, model_f = make_field_1p1(seg)

            if seg == "SC/EV":
                has_ev = True
                ev_cd2_rr = e["cd2_1p1"]
                ev_cd2_tp = e["cd2_satp"]
                # NOTE: removed a stale "fuel = FUEL_PETROL" override here.
                # veh_info_1p1() already returns FUEL_ALL for "SC/EV" (fixed
                # earlier), but this line was clobbering it back to
                # Petrol-only, silently undoing that fix for the main SC/EV
                # row. The dedicated Bike+Electric bonus row below (using
                # the same cd2_1p1/cd2_satp rate) is unaffected either way.

            pt1, pp1 = resolve_payin(e["cd2_1p1"])
            pts, pps = resolve_payin(e["cd2_satp"])

            name_rr = f"{pfx}_all_RR_{counter}_TW"
            rr_rows.append(build_row(name_rr, state, rto_f,
                                     "Comprehensive", "renew, rollover", "ALL",
                                     vt, fuel, mf, model_f,
                                     ccf, cct, None, None,
                                     pt1, pp1, eff_s, eff_e))
            counter += 1

            name_tp = f"{pfx}_all_TP_{counter}_TW"
            tp_rows.append(build_row(name_tp, state, rto_f,
                                     "TP", "ALL", "ALL",
                                     vt, fuel, mf, model_f,
                                     ccf, cct, None, None,
                                     pts, pps, eff_s, eff_e))
            counter += 1

        rows.extend(rr_rows)
        rows.extend(tp_rows)

        if has_ev:
            pt1, pp1 = resolve_payin(ev_cd2_rr)
            pts, pps = resolve_payin(ev_cd2_tp)
            rows.append(build_row(f"{pfx}_ALL_EV_RR", state, rto_f,
                                  "Comprehensive", "renew, rollover", "ALL",
                                  "Bike", FUEL_EV, "ALL", "ALL",
                                  None, None, None, None,
                                  pt1, pp1, eff_s, eff_e))
            rows.append(build_row(f"{pfx}_TP_EV_1", state, rto_f,
                                  "TP", "ALL", "ALL",
                                  "Bike", FUEL_EV, "ALL", "ALL",
                                  None, None, None, None,
                                  pts, pps, eff_s, eff_e))

    return rows, counter


def gen_saod(d, eff_s, eff_e, counter, state_filter=None):
    rows = []
    clusters = {}
    for e in d["tw_saod"]:
        clusters.setdefault(e["cluster"], []).append(e)

    yr_keys = ["yr1", "yr2", "yr3", "yr4"]
    yr_ages = ["1", "2", "3", "4"]

    for cluster in sorted(clusters):
        entries = clusters[cluster]
        rtos = sorted(d["c_saod"].get(cluster.strip().upper(), []))
        if not rtos:
            continue
        state = _state_from_rtos(rtos)
        if state_filter and state_filter not in state.split(", "):
            continue
        rto_f = _rto_field(rtos)
        pfx = _abbr(cluster)

        for yr_key, yr_age in zip(yr_keys, yr_ages):
            for e in entries:
                vt, fuel, ccf, cct, pwf, pwt, mf = veh_info_saod(e["seg"])
                pt, pp = resolve_payin(e[yr_key])
                name = f"{pfx}_all_Od_{counter}_TW"
                rows.append(build_row(name, state, rto_f,
                                      "SAOD", "renew, rollover", yr_age,
                                      vt, fuel, mf, "ALL",
                                      ccf, cct, pwf, pwt,
                                      pt, pp, eff_s, eff_e))
                counter += 1

    return rows, counter

# ── Output writer ─────────────────────────────────────────────

def write_output_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OUTPUT_HEADERS)
    hfill = PatternFill("solid", fgColor="366092")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 45)
    wb.save(path)

# ── Public API ────────────────────────────────────────────────

def generate_for_state(d, state_name, eff_s, eff_e, counter_start=1):
    rows, counter = [], counter_start
    r, counter = gen_1p5(d, eff_s, eff_e, counter, state_filter=state_name)
    rows.extend(r)
    r, counter = gen_saod(d, eff_s, eff_e, counter, state_filter=state_name)
    rows.extend(r)
    r, counter = gen_1p1(d, eff_s, eff_e, counter, state_filter=state_name)
    rows.extend(r)
    return rows, counter

def process_all(input_path, output_dir, eff_s, eff_e,
                states=None, progress_callback=None,
                output_mode="per_state", combined_filename=None):
    """
    output_mode:
      "per_state" - one workbook per state (default, original behavior)
      "combined"  - a single workbook containing every selected state's rows
      "both"      - per-state workbooks AND one combined workbook

    Rule Code numbering runs continuously across the whole run (not reset
    per state), so codes stay unique whether they land in per-state files
    or the combined file.
    """
    os.makedirs(output_dir, exist_ok=True)
    if progress_callback:
        progress_callback("Loading source data...", 0, 1)
    d = load_input_data(input_path)
    all_states = get_all_states(d)
    targets = [s for s in all_states if (states is None or s in states)]

    want_per_state = output_mode in ("per_state", "both")
    want_combined = output_mode in ("combined", "both")

    generated = []
    combined_rows = []
    counter = 1
    for idx, state in enumerate(targets):
        if progress_callback:
            progress_callback(f"Processing {state}...", idx, len(targets))
        rows, counter = generate_for_state(d, state, eff_s, eff_e, counter)
        if not rows:
            continue
        if want_per_state:
            fname = state.replace(", ", "_").replace(" ", "").replace(",", "_") + "_2W.xlsx"
            out = os.path.join(output_dir, fname)
            write_output_excel(rows, out)
            generated.append(out)
        if want_combined:
            combined_rows.extend(rows)

    if want_combined and combined_rows:
        fname = combined_filename or "Combined_States_2W.xlsx"
        if not fname.lower().endswith(".xlsx"):
            fname += ".xlsx"
        out = os.path.join(output_dir, fname)
        write_output_excel(combined_rows, out)
        generated.append(out)

    if progress_callback:
        progress_callback("Complete!", len(targets), len(targets))
    return generated

if __name__ == "__main__":
    import time
    INPUT = "/mnt/user-data/uploads/Large_Insurance_Brokers_Feb_26_h_M.xlsx"
    os.makedirs("/tmp/out_v3", exist_ok=True)
    d = load_input_data(INPUT)

    # Quick check Andaman
    rows, _ = generate_for_state(d, "ANDAMAN ISLANDS", "2026-02-01", "2026-02-28")
    print(f"Andaman: {len(rows)} rows (ref=61)")

    # Quick check AP
    rows, _ = generate_for_state(d, "ANDHRA PRADESH", "2026-02-01", "2026-02-28")
    print(f"AP: {len(rows)} rows (ref=139)")

    t0 = time.time()
    files = process_all(INPUT, "/tmp/out_v3", "2026-02-01", "2026-02-28")
    print(f"All states done in {time.time()-t0:.1f}s — {len(files)} files")
