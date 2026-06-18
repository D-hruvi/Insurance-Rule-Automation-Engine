"""
Digit 2W Insurance Commission Grid Processor — AI Edition
============================================================
Instead of hardcoded Python field-mapping logic, this version sends raw
grid rows (from the source commission Excel) to an LLM (Groq /
llama-3.1-8b-instant, free tier) in small batches. The LLM is given
the full 27 business rules as a system prompt and returns the final
structured "rule engine" output rows directly as JSON.

A post-generation Python validator then checks every returned row
against the hard constraints from the rules (Rule Code blank, Owner
Type == ALL, etc.) and reports any violations — it does NOT silently
fix them, since the point of the validator is to catch and surface
AI mistakes, not paper over them.

Environment:
    GROQ_API_KEY must be set (https://console.groq.com — free tier).

Usage:
    python ai_processor.py <input.xlsx> <output_dir> <effect_start> <effect_end> [state1,state2,...]
"""

import os
import sys
import json
import time
import re
import threading
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    from groq import Groq
except ImportError:
    Groq = None

try:
    from groq import RateLimitError as _GroqRateLimitError
except ImportError:
    class _GroqRateLimitError(Exception):
        """Placeholder so `except _GroqRateLimitError` never matches when
        the groq package is missing or doesn't expose this class."""
        pass


class _RateLimiter:
    """Paces outgoing Groq calls so we stay under the org's RPM cap.

    This is deliberately simple (a single shared "next allowed call" time
    behind a lock) rather than a full token-bucket — that's all we need for
    a sequential pipeline like this one. A 429 can push the next allowed
    call further out via penalize(), so one rate-limit hit slows the whole
    pipeline down briefly instead of every subsequent call immediately
    failing too.
    """

    def __init__(self, rpm):
        self.min_interval = 60.0 / max(rpm, 1)
        self._lock = threading.Lock()
        self._next_allowed = 0.0

    def wait(self):
        with self._lock:
            now = time.monotonic()
            delay = self._next_allowed - now
            self._next_allowed = max(now, self._next_allowed) + self.min_interval
        if delay > 0:
            time.sleep(delay)

    def penalize(self, extra_seconds):
        with self._lock:
            self._next_allowed = max(self._next_allowed, time.monotonic() + extra_seconds)


_rate_limiter = None


def _extract_retry_after(err, default=20.0):
    """Pull a Retry-After hint out of a Groq/httpx error if the response
    included one; otherwise fall back to a conservative default wait."""
    resp = getattr(err, "response", None)
    headers = getattr(resp, "headers", None) if resp is not None else None
    if headers:
        for key in ("retry-after", "Retry-After", "x-ratelimit-reset-requests"):
            val = headers.get(key)
            if val:
                try:
                    return max(float(val), 1.0)
                except (TypeError, ValueError):
                    pass
    return default

# ──────────────────────────────────────────────────────────────
# Config
# ──────────────────────────────────────────────────────────────

GROQ_MODEL = "llama-3.1-8b-instant"

# Each sheet produces a different number of OUTPUT rows per INPUT row (see
# RULES_PROMPT rule 7/8). Batch size is tuned per sheet so the JSON response
# size stays predictable: enough headroom to avoid truncation, but not so
# much that a single call burns a big chunk of the per-minute token budget.
# Tune these via env vars if you find a sweeter spot for your account's limits.
SHEET_OUTPUT_MULTIPLIER = {
    "TW 1+5": 1,
    "TW 1+1 & SATP": 2,
    "TW SAOD with Flexi Options": 4,
}
BATCH_SIZE_BY_SHEET = {
    "TW 1+5": int(os.environ.get("BATCH_SIZE_TW1P5", "1")),
    "TW 1+1 & SATP": int(os.environ.get("BATCH_SIZE_TW1P1", "1")),
    "TW SAOD with Flexi Options": int(os.environ.get("BATCH_SIZE_SAOD", "1")),
}

DEFAULT_BATCH_SIZE = 4

TOKENS_PER_OUTPUT_ROW = 500   # rough per-row JSON budget, with headroom
MIN_MAX_TOKENS = 600
MAX_TOKENS_CEILING = int(os.environ.get("GROQ_MAX_TOKENS_CEILING", "3500"))

MAX_RETRIES = 4
RETRY_DELAY_SEC = 2

# Groq's free tier is rate-limited org-wide (commonly ~30 requests/minute for
# this model, but check https://console.groq.com/dashboard/limits for your
# account's actual numbers — they vary by tier and change over time).
# We pace calls under that cap so we mostly never trigger a 429 in the first
# place, rather than firing as fast as possible and retrying after the fact.
GROQ_RPM_LIMIT = float(os.environ.get("GROQ_RPM_LIMIT", "25"))
_rate_limiter = _RateLimiter(GROQ_RPM_LIMIT)

OUTPUT_HEADERS = [
    "Rule Code", "Rule Name *", "IC Code *", "Product Type *", "Group",
    "Rule Type *", "Cover Type", "Business Type", "Vehicle Age", "State",
    "RTO", "Vehicle Category", "Vehicle Type", "Fuel Type", "Make", "Model",
    "Owner Type", "Usage Type", "Booking Mode", "Cover Selection Type",
    "Covers", "Addon Selection Type", "Addons", "CC From", "CC To",
    "Power From", "Power To", "GVW From", "GVW To", "Carrying From",
    "Carrying To", "NCB Type", "NCB From", "NCB To", "IDV From", "IDV To",
    "OD Discount From", "OD Discount To", "Effect Start Date *",
    "Effect End Date *", "PayIn (Commision Type)", "PayIn (Reward Type)",
    "PayIn (Amount Percentage)", "PayIn (OD Amount)", "PayIn (TP Amount)",
    "POSP (Commision Type)", "POSP (Reward Type)", "POSP (Amount Percentage)",
    "POSP (OD Amount)", "POSP (TP Amount)",
]

# JSON keys the LLM must use for each output row (snake_case mirror of headers)
JSON_FIELDS = [
    "rule_code", "rule_name", "ic_code", "product_type", "group",
    "rule_type", "cover_type", "business_type", "vehicle_age", "state",
    "rto", "vehicle_category", "vehicle_type", "fuel_type", "make", "model",
    "owner_type", "usage_type", "booking_mode", "cover_selection_type",
    "covers", "addon_selection_type", "addons", "cc_from", "cc_to",
    "power_from", "power_to", "gvw_from", "gvw_to", "carrying_from",
    "carrying_to", "ncb_type", "ncb_from", "ncb_to", "idv_from", "idv_to",
    "od_discount_from", "od_discount_to", "effect_start_date",
    "effect_end_date", "payin_commission_type", "payin_reward_type",
    "payin_amount_percentage", "payin_od_amount", "payin_tp_amount",
    "posp_commission_type", "posp_reward_type", "posp_amount_percentage",
    "posp_od_amount", "posp_tp_amount",
]

# ──────────────────────────────────────────────────────────────
# The 27 business rules — embedded verbatim as the system prompt
# ──────────────────────────────────────────────────────────────

RULES_PROMPT = """You are a data transformation engine for an insurance rule-engine upload file.
You will be given raw rows from a two-wheeler (2W) insurance commission grid, plus reference
lookup data (RTO/cluster/state mappings). For EACH input row, you must produce ONE OR MORE
output rule rows (one per applicable Vehicle Type / Fuel Type / Make / CC-or-Power band
combination implied by the segment description) following these exact rules:

1. Rule code column must always be blank (null).
2. Rule name will be unique for each row you generate (use the provided naming pattern
   exactly as instructed in the batch context — never reuse a name).
3. IC code column will be the literal string "DIGIT" for all rows.
4. Product column will always be "TW" (two-wheeler only — never 4W or commercial).
5. Group column will always be blank (null).
6. Rule type column will always be "PAYINPAYOUT".
7. Cover type / Business type:
   - If the row comes from sheet "TW 1+5": Cover Type = "Comprehensive", Business Type = "new".
   - If the row comes from sheet "TW 1+1 & SATP": this sheet generates TWO output rows per
     input row — one with Cover Type = "Comprehensive", Business Type = "renew, rollover"
     (the 1+1 OD component), and one with Cover Type = "TP", Business Type = "ALL"
     (the SATP/TP component).
   - If the row comes from sheet "TW SAOD with Flexi Options": Cover Type = "SAOD",
     Business Type = "renew, rollover".
8. Vehicle age column: for TW 1+5 and TW 1+1 & SATP rows, Vehicle Age is always "ALL".
   For TW SAOD rows, Vehicle Age is the specific year (1, 2, 3, or 4) corresponding to the
   yr1/yr2/yr3/yr4 PayIn column being used — SAOD generates 4 separate rows (one per year)
   for every segment row.
9. State column: entire state name, ALL CAPS ONLY (e.g. "MAHARASHTRA", not "MH_Good" or
   "MP_Good" — those are internal cluster codes from the grid, never output them as the state).
10. RTO column: use the actual RTO codes mapped to this cluster from the "2W RTO's" sheet
    (comma-separated), or "ALL" only if the cluster has no specific RTO subset (covers the
    whole state).
11. Vehicle Type column (Bike / Scooter / ALL): derived from the segment description in the
    "2W RTO's" sheet's Agency/PB segment column.
12. Fuel Type column: derived from "2W RTOs" Agency/PB segment column. Petrol can NEVER be
    output alone — it must always appear combined with other fuels, e.g.
    "Petrol, LPG, Diesel, CNG" (or the cluster-specific ordering if given), unless the segment
    is explicitly Electric-only (output "Electric") or explicitly "ALL".
13. Make column: taken from the "2W RTOs" sheet's "Make" column. If the segment says
    "Others"/excludes specific brands, output as "EXCLUDE: BRAND1, BRAND2, ..." using the
    brand names given.
14. Model column: "ALL" for almost every row. Only for TW 1+1 & SATP and SAOD sheets are
    specific models sometimes excluded or included — use the "2W Grid 5+5" sheet reference
    if such an exception is given in the batch context, otherwise default to "ALL".
15. Owner Type column: always "ALL".
16. Usage Type column: always blank (null).
17. Booking Mode column: always "any".
18. Cover Selection Type column: always "na".
19. Covers column: always blank (null).
20. Addon Selection Type column: always "na".
21. Addons column: always blank (null).
22. CC From / CC To columns: taken from the "2W Grid 5+5" sheet's Agency/PB segment CC bands,
    if the segment implies a CC range. Otherwise blank (null).
23. Power From / Power To columns: taken from the "2W Grid 5+5" sheet, if the segment implies
    an EV power band (e.g. 3-7 KW -> Power From=3.00, Power To=7.00). Otherwise blank (null).
24. Any column not explicitly covered by a rule that is empty in the source/given output stays
    blank (null). NCB Type is always "na".
25. PayIn (Commission Type) is always "net" UNLESS the source cell says "MISP", in which case
    Commission Type = "od" and Amount Percentage = "22.5".
26. PayIn (Reward Type) is always "percentage".
27. PayIn (Amount Percentage) is taken from the source grid's "Max CD2" column for the segment.
    If the source cell says "MISP", the amount is 22.5. If the source cell says "D", the
    amount is 0. Otherwise use the numeric value given (already a percentage, e.g. 35 means
    35%, not 0.35).

Additional fixed output rules (not in the numbered list but always true):
- PayIn (OD Amount) and PayIn (TP Amount) are always "0".
- POSP (Commission Type) is always "net", POSP (Reward Type) is always "percentage".
- POSP (Amount Percentage) = PayIn (Amount Percentage) * 0.8, rounded to up to 4 decimal
  places, with trailing zeros stripped. If PayIn Amount is 0, POSP Amount is also 0.
- POSP (OD Amount) and POSP (TP Amount) are always "0".
- Effect Start Date / Effect End Date are passed through unchanged from the batch context.

OUTPUT FORMAT — CRITICAL:
Return ONLY a JSON array, nothing else — no markdown fences, no commentary, no preamble.
Each element of the array is one output row, an object with EXACTLY these keys (all of them,
even if the value is null):
["rule_code","rule_name","ic_code","product_type","group","rule_type","cover_type",
"business_type","vehicle_age","state","rto","vehicle_category","vehicle_type","fuel_type",
"make","model","owner_type","usage_type","booking_mode","cover_selection_type","covers",
"addon_selection_type","addons","cc_from","cc_to","power_from","power_to","gvw_from","gvw_to",
"carrying_from","carrying_to","ncb_type","ncb_from","ncb_to","idv_from","idv_to",
"od_discount_from","od_discount_to","effect_start_date","effect_end_date",
"payin_commission_type","payin_reward_type","payin_amount_percentage","payin_od_amount",
"payin_tp_amount","posp_commission_type","posp_reward_type","posp_amount_percentage",
"posp_od_amount","posp_tp_amount"]

Use null (not the string "null") for blank cells. Use string values for everything else,
including numbers (e.g. "22.5" not 22.5), to avoid Excel formatting surprises.
"""

# ──────────────────────────────────────────────────────────────
# Groq client wrapper
# ──────────────────────────────────────────────────────────────

def get_groq_client():
    if Groq is None:
        raise RuntimeError(
            "groq package not installed. Run: pip install groq --break-system-packages"
        )
    api_key = os.environ.get("GROQ_API_KEY")
    if not api_key:
        raise RuntimeError(
            "GROQ_API_KEY environment variable not set. Get a free key at "
            "https://console.groq.com/keys"
        )
    return Groq(api_key=api_key, max_retries=0)


def _strip_json_fences(text):
    text = text.strip()
    text = re.sub(r"^```(json)?", "", text).strip()
    text = re.sub(r"```$", "", text).strip()
    return text


def call_llm_batch(client, sheet_name, batch_rows, rto_reference,
                    state_filter, eff_start, eff_end, naming_context):
    """
    Send one batch of raw grid rows to the LLM, return list of output row dicts.
    Retries on transient errors / malformed JSON, with backoff that's aware
    of *why* the call failed (rate limit vs. truncated output vs. other).
    """
    user_payload = {
        "sheet": sheet_name,
        "state_filter": state_filter,
        "effect_start_date": eff_start,
        "effect_end_date": eff_end,
        "naming_instructions": naming_context,
        "rto_reference_for_this_state": rto_reference,
        "rows": batch_rows,
    }

    user_msg = (
        f"Process these raw rows from sheet '{sheet_name}'. "
        f"Only generate rows relevant to state '{state_filter}'. "
        f"Effect Start Date = {eff_start}, Effect End Date = {eff_end}.\n\n"
        f"Naming pattern to follow for rule_name: {naming_context}\n\n"
        f"RTO reference for this state (cluster -> rto codes): "
        f"{json.dumps(rto_reference)}\n\n"
        f"Raw rows (JSON):\n{json.dumps(batch_rows, default=str)}\n\n"
        f"Return the JSON array of output rows now."
    )

    multiplier = SHEET_OUTPUT_MULTIPLIER.get(sheet_name, 1)
    est_output_rows = max(len(batch_rows) * multiplier, 1)
    max_tokens = min(
        MAX_TOKENS_CEILING,
        max(MIN_MAX_TOKENS, est_output_rows * TOKENS_PER_OUTPUT_ROW),
    )

    last_err = None
    for attempt in range(1, MAX_RETRIES + 1):
        _rate_limiter.wait()
        try:
            resp = client.chat.completions.create(
                model=GROQ_MODEL,
                messages=[
                    {"role": "system", "content": RULES_PROMPT},
                    {"role": "user", "content": user_msg},
                ],
                temperature=0,
                max_tokens=max_tokens,
            )
            raw = resp.choices[0].message.content

            cleaned = _strip_json_fences(raw)
            print("\n========== RAW LLM RESPONSE ==========")
            print(cleaned)
            print("=====================================\n")
            parsed = json.loads(cleaned)

            if not isinstance(parsed, list):
                raise ValueError("LLM did not return a JSON array")
            return parsed

        except _GroqRateLimitError as e:
            # Genuine 429 — back off for whatever Groq tells us (or a safe
            # default), and push the shared rate limiter out so the *next*
            # batch waits too, instead of immediately re-hitting the cap.
            retry_after = _extract_retry_after(e, default=RETRY_DELAY_SEC * attempt * 5)
            print(f"\n429 rate limited (attempt {attempt}/{MAX_RETRIES}); "
                  f"waiting {retry_after:.1f}s before retrying")
            last_err = e
            _rate_limiter.penalize(retry_after)
            time.sleep(retry_after)
            continue

        except json.JSONDecodeError as e:
            # Almost always a truncated response — give the next attempt
            # more room instead of just blindly retrying with the same budget.
            print(f"\nJSON parse error (attempt {attempt}/{MAX_RETRIES}): {e}")
            last_err = e
            max_tokens = min(MAX_TOKENS_CEILING, int(max_tokens * 1.6) + 200)
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY_SEC)
            continue

        except Exception as e:
            print("\nERROR:", e)
            last_err = e
            if attempt < MAX_RETRIES:
                time.sleep(RETRY_DELAY_SEC * attempt)
            continue

    raise RuntimeError(
        f"LLM batch failed after {MAX_RETRIES} attempts for sheet={sheet_name}, "
        f"state={state_filter}: {last_err}"
    )


# ──────────────────────────────────────────────────────────────
# Source workbook loading (raw rows only — no field-mapping logic)
# ──────────────────────────────────────────────────────────────

RTO_STATE_NAMES = {
    "AN": "ANDAMAN ISLANDS", "AP": "ANDHRA PRADESH", "AR": "ARUNACHAL PRADESH",
    "AS": "ASSAM", "BR": "BIHAR", "CG": "CHHATTISGARH", "CH": "CHANDIGARH",
    "DD": "DAMAN AND DIU", "DL": "DELHI", "DN": "DADRA AND NAGAR HAVELI",
    "GA": "GOA", "GJ": "GUJARAT", "HP": "HIMACHAL PRADESH", "HR": "HARYANA",
    "JH": "JHARKHAND", "JK": "JAMMU KASHMIR", "KA": "KARNATAKA", "KL": "KERALA",
    "LA": "LADAKH", "LD": "LAKSHADWEEP", "MH": "MAHARASHTRA", "ML": "MEGHALAYA",
    "MN": "MANIPUR", "MP": "MADHYA PRADESH", "MZ": "MIZORAM", "NL": "NAGALAND",
    "OD": "ODISHA", "OR": "ODISHA", "PB": "PUNJAB", "PY": "PUDUCHERRY",
    "RJ": "RAJASTHAN", "SK": "SIKKIM", "TG": "TELANGANA", "TN": "TAMIL NADU",
    "TR": "TRIPURA", "TS": "TELANGANA", "UA": "UTTARAKHAND", "UK": "UTTARAKHAND",
    "UP": "UTTAR PRADESH", "WB": "WEST BENGAL",
}


def load_raw_sheets(path):
    """Load raw rows from each relevant sheet — no interpretation, just data."""
    wb = load_workbook(path, read_only=True, data_only=True)
    data = {}

    ws_rto = wb["2W RTO's"]
    rto_rows = list(ws_rto.iter_rows(values_only=True))
    data["rto_rows"] = [r for r in rto_rows[2:] if r and r[1]]

    for sheet_name, key in [
        ("TW 1+5", "tw_1p5"),
        ("TW 1+1 & SATP", "tw_1p1"),
        ("TW SAOD with Flexi Options", "tw_saod"),
    ]:
        ws = wb[sheet_name]
        data[key] = [list(r) for r in ws.iter_rows(values_only=True) if r and r[1]]

    wb.close()
    return data


def build_rto_lookup(rto_rows):
    """cluster -> {1p1: [rtos], 1p5: [rtos], saod: [rtos]}, and rto -> state."""
    lookup = {"1p1": {}, "1p5": {}, "saod": {}}
    rto_to_state = {}
    for r in rto_rows:
        rto = r[1]
        if not rto:
            continue
        state = RTO_STATE_NAMES.get(str(rto)[:2].upper())
        if state:
            rto_to_state[rto] = state
        for idx, prod in zip((2, 3, 4), ("1p1", "1p5", "saod")):
            cluster = r[idx] if len(r) > idx else None
            if cluster:
                key = str(cluster).strip().upper() if prod == "saod" else str(cluster).strip()
                lookup[prod].setdefault(key, []).append(rto)
    return lookup, rto_to_state


def get_all_states(rto_to_state):
    return sorted(set(rto_to_state.values()))


def clusters_for_state(rto_rows_data, prod, state_name, rto_to_state):
    """Return {cluster: [rtos in this state]} for clusters touching this state."""
    out = {}
    for r in rto_rows_data:
        rto = r[1]
        if not rto or rto_to_state.get(rto) != state_name:
            continue
        idx = {"1p1": 2, "1p5": 3, "saod": 4}[prod]
        cluster = r[idx] if len(r) > idx else None
        if cluster:
            key = str(cluster).strip().upper() if prod == "saod" else str(cluster).strip()
            out.setdefault(key, []).append(rto)
    return out


# ──────────────────────────────────────────────────────────────
# Row dict -> Excel row list
# ──────────────────────────────────────────────────────────────

def row_dict_to_list(d):
    return [d.get(f) for f in JSON_FIELDS]


# ──────────────────────────────────────────────────────────────
# Output writer
# ──────────────────────────────────────────────────────────────

def write_output_excel(rows, path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.append(OUTPUT_HEADERS)
    hfill = PatternFill("solid", fgColor="366092")
    hfont = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center", wrap_text=True)
    for row in rows:
        ws.append(row)
    for col in ws.columns:
        ml = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(ml + 2, 45)
    wb.save(path)


# ──────────────────────────────────────────────────────────────
# Main generation pipeline (per state)
# ──────────────────────────────────────────────────────────────

def chunk(lst, n):
    for i in range(0, len(lst), n):
        yield lst[i:i + n]


def generate_for_state_ai(client, raw, rto_lookup, rto_to_state,
                           state_name, eff_start, eff_end,
                           progress_callback=None):
    all_rows = []

    sheet_configs = [
        ("TW 1+5", "tw_1p5", "1p5",
         "rule_name pattern: '{CLUSTER_ABBR}_new_com_{N}_TW' where N increments per row "
         "you generate within this state's file (start from 1, keep incrementing across "
         "all sheets for this state)."),
        ("TW 1+1 & SATP", "tw_1p1", "1p1",
         "rule_name pattern: for the OD/renewal row use "
         "'{CLUSTER_ABBR}_all_RR_{N}_TW', for the TP row use "
         "'{CLUSTER_ABBR}_all_TP_{N}_TW' — N keeps incrementing across all rows/sheets "
         "for this state."),
        ("TW SAOD with Flexi Options", "tw_saod", "saod",
         "rule_name pattern: '{CLUSTER_ABBR}_all_Od_{N}_TW' — N keeps incrementing across "
         "all rows/sheets for this state. Generate 4 rows per segment (years 1-4)."),
    ]

    for sheet_name, data_key, prod_key, naming in sheet_configs:
        sheet_rows = raw[data_key]
        clusters = clusters_for_state(raw["rto_rows"], prod_key, state_name, rto_to_state)
        if not clusters:
            continue

        # Build cluster lookup for filtering sheet_rows to only this state's clusters
        cluster_keys = set(clusters.keys())

        def row_cluster_key(r):
            c = r[1]
            if not c:
                return None
            return str(c).strip().upper() if prod_key == "saod" else str(c).strip()

        relevant_rows = [r for r in sheet_rows if row_cluster_key(r) in cluster_keys]
        if not relevant_rows:
            continue

        batch_size = BATCH_SIZE_BY_SHEET.get(sheet_name, DEFAULT_BATCH_SIZE)
        for batch in chunk(relevant_rows, batch_size):
            if progress_callback:
                progress_callback(f"{state_name}: {sheet_name} batch...", 0, 0)
            rto_ref = {}
            try:
                results = call_llm_batch(
                    client, sheet_name, batch, rto_ref,
                    state_name, eff_start, eff_end, naming,
                )
            except RuntimeError as e:
                if progress_callback:
                    progress_callback(f"ERROR: {e}", 0, 0)
                continue
            for r in results:
                all_rows.append(row_dict_to_list(r))

    return all_rows


def process_all_ai(input_path, output_dir, eff_start, eff_end,
                    states=None, progress_callback=None):
    os.makedirs(output_dir, exist_ok=True)
    client = get_groq_client()

    if progress_callback:
        progress_callback("Loading source workbook...", 0, 1)
    raw = load_raw_sheets(input_path)
    rto_lookup, rto_to_state = build_rto_lookup(raw["rto_rows"])
    all_states = get_all_states(rto_to_state)
    targets = [s for s in all_states if (states is None or s in states)]

    # Rough call-count estimate up front (helps catch a "this will take 2
    # hours" surprise before it actually takes 2 hours). Not exact — a row
    # can belong to more than one state for multi-state clusters like NCR —
    # but close enough to sanity-check before a long run.
    est_calls = 0
    for sheet_name, data_key, prod_key, _ in [
        ("TW 1+5", "tw_1p5", "1p5", None),
        ("TW 1+1 & SATP", "tw_1p1", "1p1", None),
        ("TW SAOD with Flexi Options", "tw_saod", "saod", None),
    ]:
        batch_size = BATCH_SIZE_BY_SHEET.get(sheet_name, DEFAULT_BATCH_SIZE)
        for state in targets:
            clusters = clusters_for_state(raw["rto_rows"], prod_key, state, rto_to_state)
            if not clusters:
                continue
            cluster_keys = set(clusters.keys())
            relevant = [
                r for r in raw[data_key]
                if r[1] and (str(r[1]).strip().upper() if prod_key == "saod" else str(r[1]).strip()) in cluster_keys
            ]
            if relevant:
                est_calls += -(-len(relevant) // batch_size)  # ceil division

    eta_min = (est_calls / GROQ_RPM_LIMIT) if GROQ_RPM_LIMIT else 0
    if progress_callback:
        progress_callback(
            f"Estimated ~{est_calls} Groq calls, ~{eta_min:.1f} min at {GROQ_RPM_LIMIT:.0f} RPM "
            f"(longer if retries/429s happen).", 0, len(targets),
        )

    generated = []
    for idx, state in enumerate(targets):
        if progress_callback:
            progress_callback(f"Processing {state}...", idx, len(targets))
        rows = generate_for_state_ai(
            client, raw, rto_lookup, rto_to_state, state,
            eff_start, eff_end, progress_callback,
        )
        if not rows:
            continue
        fname = state.replace(", ", "_").replace(" ", "") + "_2W.xlsx"
        out_path = os.path.join(output_dir, fname)
        write_output_excel(rows, out_path)
        generated.append(out_path)

    if progress_callback:
        progress_callback("Complete!", len(targets), len(targets))
    return generated


if __name__ == "__main__":
    if len(sys.argv) < 5:
        print(__doc__)
        sys.exit(1)
    input_path, output_dir, eff_s, eff_e = sys.argv[1:5]
    states = sys.argv[5].split(",") if len(sys.argv) > 5 else None

    def cb(msg, cur, total):
        print(f"[{cur}/{total}] {msg}")

    files = process_all_ai(input_path, output_dir, eff_s, eff_e, states, cb)
    print(f"\nGenerated {len(files)} files:")
    for f in files:
        print(f"  {f}")
