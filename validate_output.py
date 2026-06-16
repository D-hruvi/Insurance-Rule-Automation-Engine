"""
Validator for AI-generated Digit 2W rule engine output.

Checks every row of a generated output file (or list of row dicts) against
the HARD constraints derivable from the 27 business rules. This does NOT
fix anything — it only reports violations so you can see where the LLM
deviated from the rules and decide whether to re-run, fix manually, or
adjust the prompt.

Usage:
    python validate_output.py <output_file.xlsx>
    python validate_output.py <output_dir>          # validates every .xlsx in dir

Exit code: 0 if no violations found in any file, 1 if any violations found.
"""

import sys
import os
import glob
from openpyxl import load_workbook

# Column indices (0-based) matching OUTPUT_HEADERS in ai_processor.py
COL = {
    "rule_code": 0, "rule_name": 1, "ic_code": 2, "product_type": 3, "group": 4,
    "rule_type": 5, "cover_type": 6, "business_type": 7, "vehicle_age": 8, "state": 9,
    "rto": 10, "vehicle_category": 11, "vehicle_type": 12, "fuel_type": 13, "make": 14,
    "model": 15, "owner_type": 16, "usage_type": 17, "booking_mode": 18,
    "cover_selection_type": 19, "covers": 20, "addon_selection_type": 21, "addons": 22,
    "ncb_type": 31,
    "payin_commission_type": 40, "payin_reward_type": 41, "payin_amount_pct": 42,
    "payin_od_amount": 43, "payin_tp_amount": 44,
    "posp_commission_type": 45, "posp_reward_type": 46, "posp_amount_pct": 47,
    "posp_od_amount": 48, "posp_tp_amount": 49,
}


def _is_blank(v):
    return v is None or v == "" or str(v).strip().lower() == "null"


def validate_row(row, row_num, seen_rule_names):
    """Return list of violation strings for a single row."""
    errs = []

    def get(field):
        idx = COL[field]
        return row[idx] if idx < len(row) else None

    # Rule 1: Rule code must always be blank
    if not _is_blank(get("rule_code")):
        errs.append(f"Row {row_num}: Rule Code should be blank, got {get('rule_code')!r}")

    # Rule 2: Rule name unique
    name = get("rule_name")
    if _is_blank(name):
        errs.append(f"Row {row_num}: Rule Name is blank")
    elif name in seen_rule_names:
        errs.append(f"Row {row_num}: Rule Name '{name}' is duplicated")
    else:
        seen_rule_names.add(name)

    # Rule 3: IC code = DIGIT
    if str(get("ic_code")).strip().upper() != "DIGIT":
        errs.append(f"Row {row_num}: IC Code should be 'DIGIT', got {get('ic_code')!r}")

    # Rule 4: Product type = TW
    if str(get("product_type")).strip().upper() != "TW":
        errs.append(f"Row {row_num}: Product Type should be 'TW', got {get('product_type')!r}")

    # Rule 5: Group blank
    if not _is_blank(get("group")):
        errs.append(f"Row {row_num}: Group should be blank, got {get('group')!r}")

    # Rule 6: Rule type = PAYINPAYOUT
    if str(get("rule_type")).strip().upper() != "PAYINPAYOUT":
        errs.append(f"Row {row_num}: Rule Type should be 'PAYINPAYOUT', got {get('rule_type')!r}")

    # Rule 7: Cover/Business type combinations
    cover = str(get("cover_type") or "").strip()
    biz = str(get("business_type") or "").strip()
    valid_combos = {
        ("Comprehensive", "new"),
        ("Comprehensive", "renew, rollover"),
        ("TP", "ALL"),
        ("SAOD", "renew, rollover"),
    }
    if (cover, biz) not in valid_combos:
        errs.append(
            f"Row {row_num}: Cover Type/Business Type combo ({cover!r}, {biz!r}) "
            f"is not one of the expected combinations"
        )

    # Rule 8: Vehicle age
    vage = str(get("vehicle_age") or "").strip()
    if cover == "SAOD":
        if vage not in ("1", "2", "3", "4"):
            errs.append(f"Row {row_num}: SAOD Vehicle Age should be 1-4, got {vage!r}")
    else:
        if vage != "ALL":
            errs.append(f"Row {row_num}: Vehicle Age should be 'ALL', got {vage!r}")

    # Rule 9: State in all caps, full name (no underscores typical of cluster codes)
    state = str(get("state") or "")
    if state != state.upper():
        errs.append(f"Row {row_num}: State should be ALL CAPS, got {state!r}")
    if "_" in state:
        errs.append(f"Row {row_num}: State looks like a cluster code, not a state name: {state!r}")

    # Rule 12: Fuel type cannot be bare "Petrol"
    fuel = str(get("fuel_type") or "").strip()
    if fuel == "Petrol":
        errs.append(f"Row {row_num}: Fuel Type cannot be bare 'Petrol' (rule 12)")

    # Rule 15: Owner type = ALL
    if str(get("owner_type") or "").strip().upper() != "ALL":
        errs.append(f"Row {row_num}: Owner Type should be 'ALL', got {get('owner_type')!r}")

    # Rule 16: Usage type blank
    if not _is_blank(get("usage_type")):
        errs.append(f"Row {row_num}: Usage Type should be blank, got {get('usage_type')!r}")

    # Rule 17: Booking mode = any
    if str(get("booking_mode") or "").strip().lower() != "any":
        errs.append(f"Row {row_num}: Booking Mode should be 'any', got {get('booking_mode')!r}")

    # Rule 18: Cover selection type = na
    if str(get("cover_selection_type") or "").strip().lower() != "na":
        errs.append(
            f"Row {row_num}: Cover Selection Type should be 'na', "
            f"got {get('cover_selection_type')!r}"
        )

    # Rule 19: Covers blank
    if not _is_blank(get("covers")):
        errs.append(f"Row {row_num}: Covers should be blank, got {get('covers')!r}")

    # Rule 20: Addon selection type = na
    if str(get("addon_selection_type") or "").strip().lower() != "na":
        errs.append(
            f"Row {row_num}: Addon Selection Type should be 'na', "
            f"got {get('addon_selection_type')!r}"
        )

    # Rule 21: Addons blank
    if not _is_blank(get("addons")):
        errs.append(f"Row {row_num}: Addons should be blank, got {get('addons')!r}")

    # Rule 24: NCB type = na
    if str(get("ncb_type") or "").strip().lower() != "na":
        errs.append(f"Row {row_num}: NCB Type should be 'na', got {get('ncb_type')!r}")

    # Rule 25/27: PayIn commission type & MISP rate check
    payin_ct = str(get("payin_commission_type") or "").strip().lower()
    payin_pct = get("payin_amount_pct")
    if payin_ct not in ("net", "od"):
        errs.append(
            f"Row {row_num}: PayIn Commission Type should be 'net' or 'od', "
            f"got {get('payin_commission_type')!r}"
        )
    if payin_ct == "od":
        try:
            if abs(float(payin_pct) - 22.5) > 0.001:
                errs.append(
                    f"Row {row_num}: PayIn Commission Type is 'od' (MISP) but "
                    f"Amount Percentage is {payin_pct!r}, expected 22.5"
                )
        except (TypeError, ValueError):
            errs.append(
                f"Row {row_num}: PayIn Commission Type is 'od' but Amount Percentage "
                f"{payin_pct!r} is not a valid number"
            )

    # Rule 26: PayIn reward type = percentage
    if str(get("payin_reward_type") or "").strip().lower() != "percentage":
        errs.append(
            f"Row {row_num}: PayIn Reward Type should be 'percentage', "
            f"got {get('payin_reward_type')!r}"
        )

    # Fixed: PayIn OD/TP amount = 0
    for f in ("payin_od_amount", "payin_tp_amount"):
        v = get(f)
        if str(v).strip() not in ("0", "0.0"):
            errs.append(f"Row {row_num}: {f} should be '0', got {v!r}")

    # Fixed: POSP commission type = net, reward type = percentage
    if str(get("posp_commission_type") or "").strip().lower() != "net":
        errs.append(
            f"Row {row_num}: POSP Commission Type should be 'net', "
            f"got {get('posp_commission_type')!r}"
        )
    if str(get("posp_reward_type") or "").strip().lower() != "percentage":
        errs.append(
            f"Row {row_num}: POSP Reward Type should be 'percentage', "
            f"got {get('posp_reward_type')!r}"
        )

    # Fixed: POSP OD/TP amount = 0
    for f in ("posp_od_amount", "posp_tp_amount"):
        v = get(f)
        if str(v).strip() not in ("0", "0.0"):
            errs.append(f"Row {row_num}: {f} should be '0', got {v!r}")

    # POSP amount = PayIn amount * 0.8
    posp_pct = get("posp_amount_pct")
    try:
        p_in = float(payin_pct) if not _is_blank(payin_pct) else 0.0
        p_out = float(posp_pct) if not _is_blank(posp_pct) else 0.0
        expected = round(p_in * 0.8, 4)
        if p_in == 0:
            if p_out != 0:
                errs.append(
                    f"Row {row_num}: PayIn is 0 so POSP should be 0, got {posp_pct!r}"
                )
        elif abs(p_out - expected) > 0.01:
            errs.append(
                f"Row {row_num}: POSP Amount {posp_pct!r} != PayIn({p_in}) * 0.8 "
                f"= {expected}"
            )
    except (TypeError, ValueError):
        errs.append(
            f"Row {row_num}: PayIn/POSP amount percentages not valid numbers "
            f"({payin_pct!r}, {posp_pct!r})"
        )

    return errs


def validate_file(path):
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()

    if not rows:
        return [f"{os.path.basename(path)}: file is empty"]

    data_rows = rows[1:]  # skip header
    violations = []
    seen_names = set()
    for i, row in enumerate(data_rows, start=2):  # Excel row numbers (1=header)
        violations.extend(validate_row(row, i, seen_names))

    return violations


def main():
    if len(sys.argv) < 2:
        print(__doc__)
        sys.exit(1)

    target = sys.argv[1]
    if os.path.isdir(target):
        files = sorted(glob.glob(os.path.join(target, "*.xlsx")))
    else:
        files = [target]

    total_violations = 0
    for f in files:
        violations = validate_file(f)
        fname = os.path.basename(f)
        if violations:
            print(f"\n{'='*60}")
            print(f"{fname}: {len(violations)} violation(s)")
            print("=" * 60)
            for v in violations:
                print(f"  ✗ {v}")
            total_violations += len(violations)
        else:
            print(f"✓ {fname}: no violations")

    print(f"\n{'='*60}")
    print(f"TOTAL: {total_violations} violation(s) across {len(files)} file(s)")
    sys.exit(1 if total_violations else 0)


if __name__ == "__main__":
    main()
